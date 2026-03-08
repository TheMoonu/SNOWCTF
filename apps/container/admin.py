from django.contrib import admin
from django.urls import path, reverse
from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import HttpResponse
from container.models import (
    DockerEngine, UserContainer, StaticFile, DockerImage, 
    ContainerEngineConfig, NetworkTopologyConfig,
)
from datetime import datetime, timedelta
from container.forms import DockerEngineAdminForm
from django.utils.html import format_html
from django.utils import timezone
from django.utils.html import escape
from comment.models import SystemNotification
from django import forms
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
import json
import logging

logger = logging.getLogger('apps.container')



@admin.register(DockerEngine)
class DockerEngineAdmin(admin.ModelAdmin):
    form = DockerEngineAdminForm
    list_display = ('name', 'engine_type', 'host_info', 
                    'config_status', 'is_active', 'created_at')
    list_filter = ('engine_type', 'host_type', 'tls_enabled', 'is_active')
    search_fields = ('name', 'host', 'domain', 'namespace')
    readonly_fields = ('created_at', 'updated_at', 'kubeconfig_file_display', 
                       'ca_cert_display', 'client_cert_display', 'client_key_display')
    actions = ['test_connection']

    def get_fieldsets(self, request, obj=None):
        """
        根据引擎类型和主机类型动态调整字段集
        注意：创建新记录时（obj=None）必须包含所有字段，由 JavaScript 控制显示/隐藏
        """
        
        # 创建新记录：包含所有字段，由 JavaScript 动态控制
        if obj is None:
            return (
                ('基本信息', {
                    'fields': (
                        'name',
                        'engine_type',
                        'host_type',
                        'host',
                        'port',
                        'domain',
                        'is_active'
                    )
                }),
                ('K8s 配置', {
                    'fields': (
                        'kubeconfig_file',
                        'namespace',
                        'verify_ssl',
                    ),
                    'description': 'K8s资源requests策略由全局配置K8S_REQUESTS_RATIO控制，无需在此设置'
                }),
                ('DockerTLS配置', {
                    'fields': (
                        'tls_enabled',
                        'ca_cert',
                        'client_cert',
                        'client_key',
                    )
                }),
                ('安全策略', {
                    'fields': (
                        'security_level',
                        'enable_network_policy',
                        'enable_seccomp',
                        'enable_service_account',
                        'allow_privileged',
                        'allow_host_network',
                        'allow_host_pid',
                        'allow_host_ipc',
                    )
                }),
                ('高级安全配置', {
                    'fields': (
                        'drop_capabilities',
                    ),
                    'classes': ('collapse',)
                }),
                
            )
        
        # 编辑现有记录：根据引擎类型显示对应字段
        # K8s 引擎配置
        if obj.engine_type == 'KUBERNETES':
            return (
                ('基本信息', {
                    'fields': (
                        'name',
                        'engine_type',
                        'host',
                        'domain',
                        'is_active'
                    )
                }),
                ('K8s 配置', {
                    'fields': (
                        'kubeconfig_file_display',
                        'kubeconfig_file',
                        'namespace',
                        'verify_ssl',
                    ),
                    'description': 'K8s资源requests策略由全局配置K8S_REQUESTS_RATIO控制，无需在此设置'
                }),
                ('安全策略', {
                    'fields': (
                        'security_level',
                        'enable_network_policy',
                        'enable_seccomp',
                        'enable_service_account',
                        'allow_privileged',
                        'allow_host_network',
                        'allow_host_pid',
                        'allow_host_ipc',
                    )
                }),
                ('高级安全配置', {
                    'fields': (
                        'drop_capabilities',
                    ),
                    'classes': ('collapse',)
                }),
                ('时间信息', {
                    'classes': ('collapse',),
                    'fields': (
                        'created_at',
                        'updated_at',
                    )
                }),
            )
        
        # Docker 本地模式
        if obj.host_type == 'LOCAL':
            return (
                ('基本信息', {
                    'fields': (
                        'name',
                        'engine_type',
                        'host_type',
                        'host',
                        'domain',
                        'is_active'
                    )
                }),
                ('安全策略', {
                    'fields': (
                        'security_level',
                        'enable_network_policy',
                        'enable_seccomp',
                        'enable_service_account',
                        'allow_privileged',
                        'allow_host_network',
                        'allow_host_pid',
                        'allow_host_ipc',
                    )
                }),
                ('高级安全配置', {
                    'fields': (
                        'drop_capabilities',
                    ),
                    'classes': ('collapse',)
                }),
                ('时间信息', {
                    'classes': ('collapse',),
                    'fields': (
                        'created_at',
                        'updated_at',
                    )
                }),
            )
        
        # Docker 远程模式
        return (
            ('基本信息', {
                'fields': (
                    'name',
                    'engine_type',
                    'host_type',
                    'host',
                    'port',
                    'domain',
                    'is_active'
                )
            }),
            ('TLS 配置', {
                'classes': ('collapse',),
                'fields': (
                    'tls_enabled',
                    'ca_cert_display',
                    'ca_cert',
                    'client_cert_display',
                    'client_cert',
                    'client_key_display',
                    'client_key',
                )
            }),
            ('安全策略', {
                'fields': (
                    'security_level',
                    'enable_network_policy',
                    'enable_seccomp',
                    'enable_service_account',
                    'allow_privileged',
                    'allow_host_network',
                    'allow_host_pid',
                    'allow_host_ipc',
                    'drop_capabilities',
                )
            }),
            ('时间信息', {
                'classes': ('collapse',),
                'fields': (
                    'created_at',
                    'updated_at',
                )
            }),
        )

    def host_info(self, obj):
        """显示主机信息"""
        if obj.engine_type == 'KUBERNETES':
            # K8s 引擎显示节点 IP 和命名空间
            return format_html(
                '节点: {} / NS: {}',
                obj.host or 'N/A',
                obj.namespace or 'default'
            )
        elif obj.host_type == 'LOCAL':
            return '本地模式'
        else:
            # 远程 Docker 模式
            return format_html(
                '{} : {}',
                obj.host or 'N/A',
                obj.port or 'N/A'
            )
    host_info.short_description = '主机信息'

    def config_status(self, obj):
        """显示配置状态（TLS 或 K8s）"""
        if obj.engine_type == 'KUBERNETES':
            # K8s 配置状态
            if obj.kubeconfig_file and obj.kubeconfig_file.name:
                try:
                    # 尝试获取文件大小
                    size = obj.kubeconfig_file.size
                    return format_html(
                        '<span style="color: green;">✓ Kubeconfig 已上传 ({:.1f} KB)</span>',
                        size / 1024
                    )
                except Exception:
                    # 如果无法获取大小，至少显示文件名
                    return format_html(
                        '<span style="color: green;">✓ Kubeconfig: {}</span>',
                        obj.kubeconfig_file.name.split('/')[-1]
                    )
            else:
                return format_html(
                    '<span style="color: blue;">使用集群内配置</span>'
                )
        else:
            # Docker TLS 状态
            if not obj.tls_enabled:
                return format_html(
                    '<span style="color: gray;">TLS 未启用</span>'
                )
            if obj.ca_cert and obj.client_cert and obj.client_key:
                return format_html(
                    '<span style="color: green;">✓ TLS 配置完整</span>'
                )
            return format_html(
                '<span style="color: red;">✗ TLS 配置不完整</span>'
            )
    config_status.short_description = '配置状态'
    
    def kubeconfig_file_display(self, obj):
        """显示已上传的 Kubeconfig 文件信息"""
        if obj and obj.kubeconfig_file and obj.kubeconfig_file.name:
            try:
                size = obj.kubeconfig_file.size
                size_kb = size / 1024
                filename = obj.kubeconfig_file.name.split('/')[-1]
                return format_html(
                    '<div style="padding: 10px; background-color: #e8f5e9; border-left: 4px solid #4caf50; margin-bottom: 10px; width: 50%;">'
                    '<strong style="color: #2e7d32;">✓ 当前文件:</strong> {}<br>'
                    '<span style="color: #666;">大小: {:.2f} KB</span><br>'
                    '<a href="{}" target="_blank" style="color: #1976d2;">查看文件</a>'
                    '</div>',
                    filename,
                    size_kb,
                    obj.kubeconfig_file.url
                )
            except Exception as e:
                filename = obj.kubeconfig_file.name.split('/')[-1]
                return format_html(
                    '<div style="padding: 10px; background-color: #e8f5e9; border-left: 4px solid #4caf50; margin-bottom: 10px; width: 50%;">'
                    '<strong style="color: #2e7d32;">✓ 当前文件:</strong> {}'
                    '</div>',
                    filename
                )
        return format_html(
            '<div style="padding: 10px; background-color: #fff3e0; border-left: 4px solid #ff9800; margin-bottom: 10px; width: 50%;">'
            '<strong style="color: #e65100;">⚠ 未上传 Kubeconfig 文件</strong><br>'
            '<span style="color: #666; font-size: 12px;">将使用集群内配置（In-Cluster Config）</span>'
            '</div>'
        )
    kubeconfig_file_display.short_description = '当前 Kubeconfig 文件'
    
    def ca_cert_display(self, obj):
        """显示已上传的 CA 证书信息"""
        if obj and obj.ca_cert and obj.ca_cert.name:
            try:
                size = obj.ca_cert.size
                size_kb = size / 1024
                filename = obj.ca_cert.name.split('/')[-1]
                return format_html(
                    '<div style="padding: 8px; background-color: #e3f2fd; border-left: 3px solid #2196f3; margin-bottom: 8px; width: 50%;">'
                    '<strong style="color: #1565c0;">✓ 当前证书:</strong> {}<br>'
                    '<span style="color: #666; font-size: 12px;">大小: {:.2f} KB</span>'
                    '</div>',
                    filename,
                    size_kb
                )
            except Exception:
                filename = obj.ca_cert.name.split('/')[-1]
                return format_html(
                    '<div style="padding: 8px; background-color: #e3f2fd; border-left: 3px solid #2196f3; margin-bottom: 8px; width: 50%;">'
                    '<strong style="color: #1565c0;">✓ 当前证书:</strong> {}'
                    '</div>',
                    filename
                )
        return format_html('<span style="color: #999;">未上传</span>')
    ca_cert_display.short_description = '当前 CA 证书'
    
    def client_cert_display(self, obj):
        """显示已上传的客户端证书信息"""
        if obj and obj.client_cert and obj.client_cert.name:
            try:
                size = obj.client_cert.size
                size_kb = size / 1024
                filename = obj.client_cert.name.split('/')[-1]
                return format_html(
                    '<div style="padding: 8px; background-color: #e3f2fd; border-left: 3px solid #2196f3; margin-bottom: 8px; width: 50%;">'
                    '<strong style="color: #1565c0;">✓ 当前证书:</strong> {}<br>'
                    '<span style="color: #666; font-size: 12px;">大小: {:.2f} KB</span>'
                    '</div>',
                    filename,
                    size_kb
                )
            except Exception:
                filename = obj.client_cert.name.split('/')[-1]
                return format_html(
                    '<div style="padding: 8px; background-color: #e3f2fd; border-left: 3px solid #2196f3; margin-bottom: 8px;width: 50%;">'
                    '<strong style="color: #1565c0;">✓ 当前证书:</strong> {}'
                    '</div>',
                    filename
                )
        return format_html('<span style="color: #999;">未上传</span>')
    client_cert_display.short_description = '当前客户端证书'
    
    def client_key_display(self, obj):
        """显示已上传的客户端密钥信息"""
        if obj and obj.client_key and obj.client_key.name:
            try:
                size = obj.client_key.size
                size_kb = size / 1024
                filename = obj.client_key.name.split('/')[-1]
                return format_html(
                    '<div style="padding: 8px; background-color: #fce4ec; border-left: 3px solid #e91e63; margin-bottom: 8px;width: 50%;">'
                    '<strong style="color: #c2185b;">✓ 当前密钥:</strong> {}<br>'
                    '<span style="color: #666; font-size: 12px;">大小: {:.2f} KB</span>'
                    '</div>',
                    filename,
                    size_kb
                )
            except Exception:
                filename = obj.client_key.name.split('/')[-1]
                return format_html(
                    '<div style="padding: 8px; background-color: #fce4ec; border-left: 3px solid #e91e63; margin-bottom: 8px;width: 50%;">'
                    '<strong style="color: #c2185b;">✓ 当前密钥:</strong> {}'
                    '</div>',
                    filename
                )
        return format_html('<span style="color: #999;">未上传</span>')
    client_key_display.short_description = '当前客户端密钥'

    def save_model(self, request, obj, form, change):
        """保存模型时的额外处理"""
        # 检测引擎类型是否发生变化（只在切换类型时清空不相关字段）
        engine_type_changed = 'engine_type' in form.changed_data
        
        if obj.engine_type == 'KUBERNETES':
            # K8s 引擎：清除 Docker 专用字段
            obj.tls_enabled = False
            obj.port = None
            obj.host_type = 'LOCAL'  # K8s 不需要这个字段，设置默认值
            
            # 只在切换引擎类型时清空 Docker 证书字段
            if engine_type_changed:
                obj.ca_cert = None
                obj.client_cert = None
                obj.client_key = None
            
            # 设置默认命名空间
            if not obj.namespace:
                obj.namespace = 'ctf-challenges'
        else:
            # Docker 引擎：清除 K8s 专用字段
            # 只在切换引擎类型时清空 kubeconfig
            if engine_type_changed:
                obj.kubeconfig_file = None
            obj.namespace = ''
            
            # 如果禁用 TLS，清除证书字段（但要检查是否真的要清空）
            if not obj.tls_enabled:
                # 检测 TLS 是否刚被禁用（从启用变为禁用）
                tls_just_disabled = 'tls_enabled' in form.changed_data and not obj.tls_enabled
                if tls_just_disabled:
                    obj.ca_cert = None
                    obj.client_cert = None
                    obj.client_key = None
        
        super().save_model(request, obj, form, change)
    
    def test_connection(self, request, queryset):
        """测试引擎连接（自定义 Action）"""
        from django.contrib import messages
        
        for engine in queryset:
            try:
                if engine.engine_type == 'KUBERNETES':
                    # 测试 K8s 连接
                    from container.k8s_service import K8sService
                    service = K8sService(engine=engine)
                    # 尝试获取命名空间信息
                    from kubernetes import client
                    core_api = client.CoreV1Api()
                    ns = core_api.read_namespace(name=engine.namespace or 'ctf-challenges')
                    
                    self.message_user(
                        request,
                        f'✓ K8s 引擎 "{engine.name}" 连接成功！命名空间: {ns.metadata.name}',
                        level=messages.SUCCESS
                    )
                else:
                    # 测试 Docker 连接
                    from container.docker_service import DockerService
                    docker_url = engine.get_docker_url()
                    tls_config = engine.get_tls_config() if engine.tls_enabled else None
                    service = DockerService(url=docker_url, tls_config=tls_config)
                    
                    import docker
                    with docker.DockerClient(
                        base_url=docker_url,
                        tls=tls_config,
                        timeout=10
                    ) as client:
                        info = client.info()
                        version = info.get('ServerVersion', 'Unknown')
                        containers = info.get('Containers', 0)
                    
                    self.message_user(
                        request,
                        f'✓ Docker 引擎 "{engine.name}" 连接成功！版本: {version}, 容器数: {containers}',
                        level=messages.SUCCESS
                    )
            except Exception as e:
                self.message_user(
                    request,
                    f'✗ 引擎 "{engine.name}" 连接失败: {str(e)}',
                    level=messages.ERROR
                )
    
    test_connection.short_description = '测试引擎连接'
    
    class Media:
        """添加自定义 JavaScript 用于动态显示/隐藏字段和安全级别配置"""
        js = (
            'container/js/admin_docker_engine_form.js',
            'container/js/admin_k8s_security.js',
        )

@admin.register(UserContainer)
class UserContainerAdmin(admin.ModelAdmin):
    list_display = (
        'user', 'challenge_title', 'container_id_short', 'container_type',
        'status_display', 'docker_engine', 'lifetime_display',
        'created_at'
    )
    # 减少筛选项，只保留最常用的
    list_filter = (
        'status',
        'container_type', 
        'docker_engine',
        'created_at'
    )
    search_fields = ('user__username', 'container_id', 'challenge_title', 'challenge_uuid', 'competition__title')
    
    # 作为日志，大部分字段应该是只读的
    readonly_fields = (
        'user', 'challenge_uuid', 'challenge_title', 'container_id',
        'container_type', 'competition', 'docker_engine',
        'ip_address', 'domain', 'port',
        'created_at', 'expires_at', 'deleted_at', 'deleted_by',
        'lifetime_info'
    )
    
    # 性能优化设置
    list_per_page = 100  # 每页显示100条
    list_max_show_all = 500
    show_full_result_count = False
    date_hierarchy = 'created_at'  # 添加日期层次导航
    
    fieldsets = (
        ('基本信息', {
            'fields': ('user', 'challenge_uuid', 'challenge_title', 'container_id')
        }),
        ('容器分类', {
            'fields': ('container_type', 'competition')
        }),
        ('网络配置', {
            'fields': ('docker_engine', 'ip_address', 'domain', 'port')
        }),
        ('状态管理', {
            'fields': ('status', 'deleted_by', 'created_at', 'expires_at', 'deleted_at', 'lifetime_info')
        }),
    )
    
    def container_id_short(self, obj):
        """显示容器ID前12位"""
        if obj.container_id:
            return obj.container_id
        return '-'
    container_id_short.short_description = '容器ID'
    
    def status_display(self, obj):
        """带颜色的状态显示"""
        colors = {
            'RUNNING': '#28a745',   # 绿色
            'DELETED': '#dc3545',   # 红色
            'EXPIRED': '#ffc107',   # 黄色
            'FAILED': '#6c757d',    # 灰色
            'STOPPED': '#17a2b8',   # 蓝色
        }
        color = colors.get(obj.status, '#6c757d')
        return format_html(
            '<span style="color: {}; font-weight: bold;">●</span> {}',
            color,
            obj.get_status_display()
        )
    status_display.short_description = '状态'
    
    def lifetime_display(self, obj):
        """显示容器运行时长"""
        seconds = obj.get_lifetime_seconds()
        minutes = seconds / 60
        hours = minutes / 60
        
        if hours >= 1:
            return f"{hours:.1f}小时"
        elif minutes >= 1:
            return f"{minutes:.1f}分钟"
        else:
            return f"{seconds:.0f}秒"
    lifetime_display.short_description = '运行时长'
    
    def lifetime_info(self, obj):
        """详细的运行时长信息（用于详情页）"""
        if not obj.created_at:
            return format_html('<pre>容器尚未创建</pre>')
        
        seconds = obj.get_lifetime_seconds()
        hours, remainder = divmod(int(seconds), 3600)
        minutes, seconds = divmod(remainder, 60)
        
        lifetime_str = f"{hours}小时 {minutes}分钟 {seconds}秒"
        
        info = f"运行时长: {lifetime_str}\n"
        info += f"创建时间: {obj.created_at.strftime('%Y-%m-%d %H:%M:%S')}\n"
        if obj.deleted_at:
            info += f"删除时间: {obj.deleted_at.strftime('%Y-%m-%d %H:%M:%S')}\n"
            info += f"删除方式: {obj.deleted_by}\n"
        
        return format_html('<pre>{}</pre>', info)
    lifetime_info.short_description = '生命周期详情'
    
    def get_queryset(self, request):
        """优化查询，预加载关联对象"""
        qs = super().get_queryset(request)
        return qs.select_related('user', 'docker_engine')
    
    # 禁用添加功能（日志性质的数据不应该手动添加）
    def has_add_permission(self, request):
        return False
    
    # 只允许超级管理员删除
    def has_delete_permission(self, request, obj=None):
        return request.user.is_superuser
    
    actions = ['mark_as_deleted', 'mark_as_expired', 'cleanup_old_logs']
    
    def mark_as_deleted(self, request, queryset):
        """批量标记为已删除"""
        count = 0
        for obj in queryset.filter(status='RUNNING'):
            obj.mark_deleted(deleted_by='ADMIN')
            count += 1
        
        self.message_user(request, f'已将 {count} 个容器标记为已删除')
    mark_as_deleted.short_description = '标记为已删除（仅更新状态，不删除Docker容器）'
    
    def mark_as_expired(self, request, queryset):
        """批量标记为已过期"""
        count = 0
        for obj in queryset.filter(status='RUNNING'):
            obj.mark_expired()
            count += 1
        
        self.message_user(request, f'已将 {count} 个容器标记为已过期')
    mark_as_expired.short_description = '标记为已过期（仅更新状态，不删除Docker容器）'
    
    def cleanup_old_logs(self, request, queryset):
        """清理7天前的已删除/已过期/失败/停止的日志（自动清理所有符合条件的记录）"""
        
        # 计算7天前的时间
        seven_days_ago = timezone.now() - timedelta(days=7)
        
        # 直接查询数据库中所有符合条件的记录（不依赖 queryset）
        from container.models import UserContainer
        old_logs = UserContainer.objects.filter(
            status__in=['DELETED', 'EXPIRED', 'FAILED', 'STOPPED'],
            deleted_at__lt=seven_days_ago
        )
        
        count = old_logs.count()
        if count > 0:
            old_logs.delete()
            self.message_user(request, f'已清理 {count} 条7天前的容器日志记录', level='success')
        else:
            self.message_user(request, '没有符合条件的日志记录（7天前的已删除/已过期/失败/停止记录）', level='warning')
    
    cleanup_old_logs.short_description = '清理7天前的日志'

@admin.register(StaticFile)
class StaticFileAdmin(admin.ModelAdmin):
    list_display = ('name', 'file_size_display', 'download_count', 
                   'upload_time', 'author', 'review_status', 'download_button', 'reviewer', 'review_time')
    list_filter = ('review_status', 'author', 'reviewer')
    search_fields = ('name', 'description')
    readonly_fields = ('file_size', 'download_count', 'upload_time', 'review_time')
    actions = ['approve_selected', 'reject_selected']
    
    # 性能优化设置
    list_per_page = 50
    list_max_show_all = 200
    show_full_result_count = False

    def get_form(self, request, obj=None, **kwargs):
        """设置表单初始值"""
        form = super().get_form(request, obj, **kwargs)
        # 创建新题目时，默认设置作者为当前用户
        if not obj and 'author' in form.base_fields:
            form.base_fields['author'].initial = request.user
        return form
    
    def get_queryset(self, request):
        """优化查询性能"""
        qs = super().get_queryset(request)
        return qs.select_related('author', 'reviewer')
    
    fieldsets = (
        ('基本信息', {
            'fields': ('name', 'file', 'description')
        }),
        ('审核信息', {
            'fields': ('review_status', 'reviewer', 'review_time', 'review_comment')
        }),
        ('其他信息', {
            'fields': ('file_size', 'download_count', 'upload_time', 'author')
        })
    )
    
    def file_size_display(self, obj):
        """将文件大小转换为人类可读格式"""
        size = obj.file_size
        for unit in ['B', 'KB', 'MB', 'GB']:
            if size < 1024:
                return f"{size:.2f} {unit}"
            size /= 1024
        return f"{size:.2f} TB"
    file_size_display.short_description = '文件大小'
    
    def download_button(self, obj):
        """显示下载按钮（纯文本链接）"""
        if obj.file and obj.review_status == 'APPROVED':
            # 获取文件 URL
            file_url = obj.get_file_url_one()
            if file_url:
                return format_html(
                    '<a href="{}" target="_blank">下载</a>',
                    file_url
                )
        return '-'
    download_button.short_description = '下载'

    def save_model(self, request, obj, form, change):
        if not change:
            obj.author = request.user
        super().save_model(request, obj, form, change)
        
    def approve_selected(self, request, queryset):
        """批量通过审核"""
        for obj in queryset:
            # 更新审核状态
            obj.review_status = 'APPROVED'
            obj.reviewer = request.user
            obj.review_time = timezone.now()
            obj.save()
            
            # 如果不是管理员创建的文件，则发送通知
            if not (obj.author and (obj.author.is_superuser or obj.author.is_staff)):
                # 创建通知
                notification = SystemNotification.objects.create(
                    title='文件审核通过',
                    content=f'''
                        <p>您上传的文件 "{escape(obj.name)}" 已通过审核。</p>
                        <p>审核时间：{timezone.now().strftime('%Y-%m-%d %H:%M')}</p>
                    '''
                )
                notification.get_p.add(obj.author)
            
        self.message_user(request, f"已通过 {queryset.count()} 个文件的审核")
    approve_selected.short_description = "通过所选文件的审核"
    
    def reject_selected(self, request, queryset):
        """批量拒绝审核"""
        for obj in queryset:
            # 更新审核状态
            obj.review_status = 'REJECTED'
            obj.reviewer = request.user
            obj.review_time = timezone.now()
            obj.save()
            
            # 如果不是管理员创建的文件，则发送通知
            if not (obj.author and (obj.author.is_superuser or obj.author.is_staff)):
                # 创建通知
                notification = SystemNotification.objects.create(
                    title='Docker Compose配置审核未通过',
                    content=f'''
                        <p>您创建的配置 "{escape(obj.name)}" 未通过审核。备注：{escape(obj.review_comment)}</p>
                        <p>审核时间：{timezone.now().strftime('%Y-%m-%d %H:%M')}</p>
                    '''
                )
                notification.get_p.add(obj.author)
            
        self.message_user(request, f"已拒绝 {queryset.count()} 个配置的审核")
    reject_selected.short_description = "拒绝所选配置的审核"




@admin.register(DockerImage)
class DockerImageAdmin(admin.ModelAdmin):
    list_display = ('name', 'tag', 'registry', 'category',
                   'ports_display', 'resource_display', 'image_status','review_status', 'is_active', 
                   'author', 'created_at')
    list_filter = ('category', 'review_status', 
                  'is_active', 'is_pulled')
    search_fields = ('name', 'tag', 'description')
    readonly_fields = ('image_id', 'image_size', 'is_pulled', 'last_pulled',
                      'created_at', 'updated_at', 'review_time')
    actions = ['export_template_action', 'import_excel_action', 'approve_selected', 'reject_selected', 'pull_images']
    
    # 性能优化设置
    list_per_page = 50  # 每页显示50条
    list_max_show_all = 200  # 最多显示200条"显示全部"
    show_full_result_count = False  # 不显示精确的总数（大数据集时提升性能）
    
    fieldsets = (
        ('基本信息', {
            'fields': ('name', 'tag', 'registry', 'category', 'description', 'entrance','is_active')
        }),
        ('Flag 配置', {
            'fields': ('flag_inject_method', 'flag_env_name', 'flag_script'),
            'description': 'INTERNAL: 使用SNOW_FLAG; CUSTOM_ENV: 自定义环境变量名; SCRIPT: 脚本注入'
        }),
        ('端口配置', {
            'fields': ('exposed_ports',),
            'description': '多个端口用逗号分隔'
        }),
        ('资源限制', {
            'fields': ('memory_limit', 'cpu_limit'),
            'description': '留空则使用默认值（内存512MB，CPU1核）。根据题目复杂度配置：轻量级(256MB/0.5核)，中型(512MB/1核)，重型(1024MB+/2核+)'
        }),
        ('镜像状态', {
            'fields': ('is_pulled', 'image_id', 'image_size', 'last_pulled'),
            'classes': ('collapse',)
        }),
        ('审核信息', {
            'fields': ('review_status', 'reviewer', 'review_time', 'review_comment')
        }),
        ('其他信息', {
            'fields': ('author', 'created_at', 'updated_at'),
            'classes': ('collapse',)
        })
    )
    def get_form(self, request, obj=None, **kwargs):
        """设置表单初始值"""
        form = super().get_form(request, obj, **kwargs)
        # 创建新题目时，默认设置作者为当前用户
        if not obj and 'author' in form.base_fields:
            form.base_fields['author'].initial = request.user
        return form
    # SimpleUI 自定义按钮 - 链接类型
    def export_template_action(self, request, queryset):
        """下载Excel模板按钮"""
        pass
    
    export_template_action.short_description = '下载Excel模板'
    export_template_action.icon = 'el-icon-download'
    export_template_action.type = 'success'
    export_template_action.action_type = 1  # 链接跳转
    export_template_action.action_url = '/adminx/container/dockerimage/export-template/'
    
    def import_excel_action(self, request, queryset):
        """批量导入镜像按钮"""
        pass
    
    import_excel_action.short_description = '批量导入镜像'
    import_excel_action.icon = 'el-icon-upload2'
    import_excel_action.type = 'primary'
    import_excel_action.action_type = 1  # 链接跳转
    import_excel_action.action_url = '/adminx/container/dockerimage/import-excel/'
    
    def get_urls(self):
        """添加自定义URL"""
        urls = super().get_urls()
        custom_urls = [
            path('import-excel/', self.admin_site.admin_view(self.import_excel_view), name='container_dockerimage_import_excel'),
            path('export-template/', self.admin_site.admin_view(self.export_template_view), name='container_dockerimage_export_template'),
        ]
        return custom_urls + urls
    

    
    def import_excel_view(self, request):
        """处理Excel导入"""
        # 权限检查
        if not self.has_add_permission(request):
            from django.core.exceptions import PermissionDenied
            raise PermissionDenied("您没有添加镜像的权限")
        
        if request.method != 'POST':
            context = dict(
                self.admin_site.each_context(request),
                title='批量导入镜像',
            )
            return render(request, 'admin/container/dockerimage/import_page.html', context)
        
        if not request.FILES.get('excel_file'):
            self.message_user(request, "请选择要导入的Excel文件", messages.ERROR)
            return redirect('.')
        
        excel_file = request.FILES['excel_file']
        
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            
            success_count = 0
            error_messages = []
            
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    if not any(row):
                        continue
                    
                    (name, tag, registry, category, description,
                     flag_inject_method, flag_env_name, flag_script,
                     exposed_ports, memory_limit, cpu_limit, is_active) = row[:12]
                    
                    if not name:
                        error_messages.append(f"第{row_num}行: 镜像名称不能为空")
                        continue
                    
                    # 处理布尔值
                    is_active_bool = str(is_active).lower() in ['true', '1', 'yes', '是', True, 1] if is_active else True
                    
                    # 创建镜像配置
                    DockerImage.objects.create(
                        name=str(name).strip(),
                        tag=str(tag or 'latest').strip(),
                        registry=str(registry or 'docker.io').strip(),
                        category=str(category or 'Web').strip(),
                        description=str(description or '').strip() if description else '',
                        flag_inject_method=str(flag_inject_method or 'INTERNAL').strip(),
                        flag_env_name=str(flag_env_name or '').strip() if flag_env_name else '',
                        flag_script=str(flag_script or '').strip() if flag_script else '',
                        exposed_ports=str(exposed_ports or '80').strip(),
                        memory_limit=int(memory_limit or 256),
                        cpu_limit=float(cpu_limit or 0.5),
                        is_active=is_active_bool,
                        review_status='PENDING',
                        author=request.user
                    )
                    
                    success_count += 1
                    
                except Exception as e:
                    error_messages.append(f"第{row_num}行: {str(e)}")
                    continue
            
            if success_count > 0:
                self.message_user(request, f"✅ 成功导入 {success_count} 个镜像配置！", messages.SUCCESS)
            else:
                self.message_user(request, "⚠️ 未导入任何数据", messages.WARNING)
            
            if error_messages:
                for msg in error_messages[:10]:
                    self.message_user(request, msg, messages.WARNING)
                if len(error_messages) > 10:
                    self.message_user(request, f"...还有 {len(error_messages) - 10} 条错误未显示", messages.WARNING)
            
        except Exception as e:
            self.message_user(request, f"❌ 导入失败: {str(e)}", messages.ERROR)
        
        from django.urls import reverse
        return redirect(reverse('admin:container_dockerimage_changelist'))
    
    def export_template_view(self, request):
        """导出Excel模板"""
        # 权限检查
        if not self.has_view_permission(request):
            from django.core.exceptions import PermissionDenied
            raise PermissionDenied("您没有查看权限")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '镜像导入模板'
        
        headers = [
            '镜像名称*', '镜像标签', '镜像仓库', '镜像类型', '描述',
            'Flag注入方式', 'Flag环境变量名', 'Flag注入脚本',
            '暴露端口', '内存限制(MB)', 'CPU限制(核)', '是否启用'
        ]
        
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        example_data = [
            'nginx',
            'latest',
            'docker.io',
            'Web',
            'Nginx Web服务器镜像',
            'INTERNAL',
            '',
            '',
            '80',
            256,
            0.5,
            'true'
        ]
        
        for col_num, value in enumerate(example_data, 1):
            cell = ws.cell(row=2, column=col_num, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        column_widths = [25, 15, 20, 15, 40, 20, 20, 40, 15, 15, 15, 12]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width
        
        ws_info = wb.create_sheet('填写说明')
        info_texts = [
            ['字段说明', ''],
            ['镜像名称*', '必填，如: nginx, mysql, redis'],
            ['镜像标签', '镜像版本标签，默认latest'],
            ['镜像仓库', '镜像仓库地址，默认docker.io'],
            ['镜像类型', f"可选：{', '.join([c[0] for c in DockerImage.CATEGORY_CHOICES[:10]])}等"],
            ['描述', '镜像的描述信息'],
            ['Flag注入方式', 'INTERNAL/CUSTOM_ENV/SCRIPT/NONE，默认INTERNAL'],
            ['Flag环境变量名', '当选择CUSTOM_ENV时填写，如: FLAG, CTF_FLAG'],
            ['Flag注入脚本', '当选择SCRIPT时填写，如: echo "$SNOW_FLAG" > /flag.txt'],
            ['暴露端口', '多个端口用逗号分隔，如: 80,3306'],
            ['内存限制(MB)', '默认256MB'],
            ['CPU限制(核)', '默认0.5核'],
            ['是否启用', 'true/false，默认true'],
            ['', ''],
            ['注意事项', ''],
            ['1', '带*号的字段为必填项'],
            ['2', '导入的镜像默认为待审核状态'],
            ['3', '布尔值字段可使用: true/false, 1/0, yes/no, 是/否'],
        ]
        
        for row_num, (field, desc) in enumerate(info_texts, 1):
            ws_info.cell(row=row_num, column=1, value=field).font = Font(bold=True)
            ws_info.cell(row=row_num, column=2, value=desc)
        
        ws_info.column_dimensions['A'].width = 20
        ws_info.column_dimensions['B'].width = 60
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="dockerimage_import_template.xlsx"'
        
        return response
    
    def flag_inject_display(self, obj):
        """显示 Flag 注入方式"""
        method_map = {
            'INTERNAL': ('标准环境变量', 'green'),
            'CUSTOM_ENV': ('自定义环境变量', 'blue'),
            'SCRIPT': ('脚本注入', 'orange'),
            'NONE': ('不支持', 'gray')
        }
        text, color = method_map.get(obj.flag_inject_method, ('未知', 'red'))
        
        # 如果是自定义环境变量，显示变量名
        if obj.flag_inject_method == 'CUSTOM_ENV' and obj.flag_env_name:
            return format_html(
                '<span style="color: {};">{}<br/><small>{}</small></span>',
                color,
                text,
                obj.flag_env_name
            )
        
        return format_html(
            '<span style="color: {};">{}</span>',
            color,
            text
        )
    flag_inject_display.short_description = 'Flag注入方式'
    
    def ports_display(self, obj):
        """显示端口信息"""
        ports = obj.get_ports_list()
        if not ports:
            return '-'
        return ', '.join(ports[:3]) + ('...' if len(ports) > 3 else '')
    ports_display.short_description = '端口'
    
    def resource_display(self, obj):
        """显示资源限制"""
        memory = obj.memory_limit or 512  # 默认512MB
        cpu = obj.cpu_limit or 1.0  # 默认1核
        
        # 根据资源大小显示不同颜色
        if memory <= 256 and cpu <= 0.5:
            color = 'green'  # 轻量级
        elif memory <= 512 and cpu <= 1.0:
            color = 'blue'  # 中型
        else:
            color = 'orange'  # 重型
        
        return format_html(
            '<span style="color: {};">{} 核 / {} MB</span>',
            color,
            cpu,
            memory
        )
    resource_display.short_description = '资源限制'
    
    def image_status(self, obj):
        """显示镜像状态（使用缓存）"""
        from django.core.cache import cache
        from container.models import DockerEngine
        
        # 获取所有激活的引擎
        engines = DockerEngine.objects.filter(is_active=True).order_by('engine_type', 'name')
        
        if not engines.exists():
            return format_html('<span style="color: gray;">无可用引擎</span>')
        
        # 为每个镜像单独缓存状态
        cache_key = f'docker_image_{obj.id}_status'
        cached_data = cache.get(cache_key)
        
        if cached_data and isinstance(cached_data, dict):
            # 使用缓存数据
            engine_statuses = cached_data.get('engine_statuses', [])
            cache_time = cached_data.get('cache_time', '')
        else:
            # 无缓存，显示未检查状态
            engine_statuses = []
            cache_time = ''
            for engine in engines:
                engine_name_display = f"{engine.name} (K8s)" if engine.engine_type == 'KUBERNETES' else engine.name
                engine_statuses.append({
                    'name': engine_name_display,
                    'status': 'unknown',
                    'color': '#999',
                    'icon': '○'
                })
        
        # 统计拉取情况
        pulled_count = sum(1 for s in engine_statuses if s['status'] == 'pulled')
        not_pulled_count = sum(1 for s in engine_statuses if s['status'] == 'not_pulled')
        unknown_count = sum(1 for s in engine_statuses if s['status'] == 'unknown')
        error_count = sum(1 for s in engine_statuses if s['status'] == 'error')
        total_count = len(engine_statuses)
        
        # 构建刷新按钮
        refresh_button = (
            f'<button type="button" class="btn-refresh-image-status" '
            f'data-image-id="{obj.id}" '
            f'style="padding: 2px 8px; font-size: 11px; cursor: pointer; '
            f'background: linear-gradient(to bottom, #e3f4ff 0%, #cfe9ff 100%); '
            f'color: #205067; border: 1px solid #b4d5e6; border-radius: 4px; '
            f'font-weight: 500; transition: all 0.2s ease; margin-left: 8px;">'
            f'🔄刷新</button>'
        )
        
        # 构建总览状态
        if unknown_count == total_count:
            # 全部未检查
            status_text = f'<strong style="color: #999;">未检查</strong>'
        elif pulled_count == total_count:
            # 全部已拉取
            status_text = f'<strong style="color: green;">✓ 已拉取 ({pulled_count}/{total_count})</strong>'
        elif pulled_count > 0:
            # 部分已拉取
            status_text = f'<strong style="color: orange;">⚠ 部分拉取 ({pulled_count}/{total_count})</strong>'
        elif error_count > 0:
            # 有错误
            status_text = f'<strong style="color: red;">✗ 检查失败 ({error_count}/{total_count})</strong>'
        else:
            # 全部未拉取
            status_text = f'<strong style="color: gray;">✗ 未拉取 (0/{total_count})</strong>'
        
        overview = (
            f'<div style="display: flex; align-items: center; margin-bottom: 5px;">'
            f'{status_text}'
            f'{refresh_button}'
            f'</div>'
        )
        
        # 开始构建完整 HTML
        html_parts = [f'<div id="image-status-{obj.id}">{overview}']
        
        # 添加每个引擎的详细状态
        for status in engine_statuses:
            name = status['name']
            color = status['color']
            icon = status['icon']
            note = status.get('note', '')
            
            if note:
                html_parts.append(
                    f'<div style="margin: 1px 0; padding-left: 8px;">'
                    f'<small style="color: {color};">{icon} {name} - <em>{note}</em></small>'
                    f'</div>'
                )
            else:
                html_parts.append(
                    f'<div style="margin: 1px 0; padding-left: 8px;">'
                    f'<small style="color: {color};">{icon} {name}</small>'
                    f'</div>'
                )
        
        # 添加缓存时间
        if cache_time:
            html_parts.append(f'<div style="margin-top: 3px;"><small style="color: #999;">更新: {cache_time}</small></div>')
        elif obj.last_pulled and pulled_count > 0:
            time_str = obj.last_pulled.strftime('%m-%d %H:%M')
            html_parts.append(f'<div style="margin-top: 3px;"><small style="color: #999;">拉取: {time_str}</small></div>')
        
        html_parts.append('</div>')
        
        return format_html(''.join(html_parts))
    
    image_status.short_description = '引擎资源'
    
    def get_queryset(self, request):
        """优化查询性能，避免N+1问题"""
        qs = super().get_queryset(request)
        # 预加载关联对象
        qs = qs.select_related('author', 'reviewer')
        return qs
    
    def save_model(self, request, obj, form, change):
        if not change:
            obj.author = request.user
        super().save_model(request, obj, form, change)
    
    def approve_selected(self, request, queryset):
        """批量通过审核"""
        for obj in queryset:
            obj.review_status = 'APPROVED'
            obj.reviewer = request.user
            obj.review_time = timezone.now()
            obj.save()
            
            # 如果不是管理员创建的，则发送通知
            if not (obj.author and (obj.author.is_superuser or obj.author.is_staff)):
                notification = SystemNotification.objects.create(
                    title='Docker 镜像配置审核通过',
                    content=f'''
                        <p>您创建的镜像配置 "{escape(obj.name)}:{escape(obj.tag)}" 已通过审核。</p>
                        <p>审核时间：{timezone.now().strftime('%Y-%m-%d %H:%M')}</p>
                    '''
                )
                notification.get_p.add(obj.author)
        
        self.message_user(request, f"已通过 {queryset.count()} 个镜像配置的审核")
    approve_selected.short_description = "通过所选镜像配置的审核"
    
    def reject_selected(self, request, queryset):
        """批量拒绝审核"""
        for obj in queryset:
            obj.review_status = 'REJECTED'
            obj.reviewer = request.user
            obj.review_time = timezone.now()
            obj.save()
            
            # 如果不是管理员创建的，则发送通知
            if not (obj.author and (obj.author.is_superuser or obj.author.is_staff)):
                notification = SystemNotification.objects.create(
                    title='Docker 镜像配置审核未通过',
                    content=f'''
                        <p>您创建的镜像配置 "{escape(obj.name)}:{escape(obj.tag)}" 未通过审核。</p>
                        <p>审核备注：{escape(obj.review_comment or '无')}</p>
                        <p>审核时间：{timezone.now().strftime('%Y-%m-%d %H:%M')}</p>
                    '''
                )
                notification.get_p.add(obj.author)
        
        self.message_user(request, f"已拒绝 {queryset.count()} 个镜像配置的审核")
    reject_selected.short_description = "拒绝所选镜像配置的审核"
    
    def pull_images(self, request, queryset):
        """异步拉取选中的镜像到所有激活的引擎"""
        import docker
        from container.models import DockerEngine
        
        # 1. 检查所有激活的引擎
        engines = DockerEngine.objects.filter(is_active=True)
        docker_engines = engines.filter(engine_type='DOCKER')
        k8s_engines = engines.filter(engine_type='KUBERNETES')
        
        # 提示引擎信息
       
        if not engines.exists():
            self.message_user(request, "没有激活的容器引擎，请先配置并激活至少一个引擎", level='error')
            return
        
        # 2. 测试所有引擎的连接
        available_engines = []
        engine_status = []
        
        for engine in engines:
            # K8s 引擎
            if engine.engine_type == 'KUBERNETES':
                try:
                    from container.k8s_service import K8sService
                    k8s_service = K8sService(engine=engine)
                    # 简单测试：列出命名空间
                    from kubernetes import client as k8s_client
                    core_api = k8s_client.CoreV1Api()
                    core_api.list_namespace(limit=1)
                    available_engines.append(engine)
                except Exception as e:
                    engine_status.append(f"{engine.name} (K8s): 连接失败 - {str(e)}")
            
            # Docker 引擎
            else:
                try:
                    docker_url = engine.get_docker_url()
                    tls_config = engine.get_tls_config() if engine.needs_tls else None
                    client = docker.DockerClient(base_url=docker_url, tls=tls_config, timeout=10)
                    
                    # 尝试 ping Docker 服务
                    client.ping()
                    
                    # 测试镜像仓库连接
                    try:
                        info = client.info()
                        if info:
                            available_engines.append(engine)
                        else:
                            engine_status.append(f"{engine.name} (Docker): 无法获取镜像仓库信息")
                    except Exception as e:
                        available_engines.append(engine)
                        engine_status.append(f"{engine.name} (Docker): 引擎可用但存在网络问题")
                    
                    client.close()
                    
                except docker.errors.DockerException as e:
                    engine_status.append(f"{engine.name} (Docker): 连接失败，请检查引擎配置")
                except Exception as e:
                    engine_status.append(f"{engine.name} (Docker): 测试失败，请检查引擎配置")
        
        # 3. 显示所有引擎的状态
        for status in engine_status:
            if status.startswith('✅'):
                self.message_user(request, status, level='success')
            elif status.startswith('⚠️'):
                self.message_user(request, status, level='warning')
            else:
                self.message_user(request, status, level='error')
        
        # 4. 检查是否有可用的引擎
        if not available_engines:
            self.message_user(
                request, 
                f"所有容器引擎都不可用（共 {engines.count()} 个），无法执行拉取任务", 
                level='error'
            )
            return
        
        # 5. 如果只有部分引擎可用，提示用户
        if len(available_engines) < engines.count():
            self.message_user(
                request,
                f"将使用 {len(available_engines)}/{engines.count()} 个可用的容器引擎执行任务",
                level='info'
            )
        
        # 6. 检查 Celery 是否可用
        try:
            from easytask.tasks import pull_multiple_docker_images
            from celery import current_app
            
            # 检查是否有活跃的 Worker
            inspect = current_app.control.inspect()
            active_workers = inspect.active()
            
            if not active_workers:
                self.message_user(
                    request, 
                    "异步任务错误", 
                    level='error'
                )
                return
                
        except ImportError:
            self.message_user(request, "Celery 未正确配置", level='error')
            return
        except Exception as e:
            self.message_user(request, f"Celery 连接失败，请检查Celery配置", level='error')
            return
        
        # 7. 所有检查通过，提交异步任务
        image_ids = list(queryset.values_list('id', flat=True))
        task = pull_multiple_docker_images.delay(image_ids)
        
        # 8. 显示友好的提示信息
        image_names = ', '.join([obj.full_name for obj in queryset[:3]])
        if len(queryset) > 3:
            image_names += f' 等共 {len(queryset)} 个镜像'
        
        # 构建引擎信息
        engine_info = f"可用引擎: {', '.join([e.name for e in available_engines])}"
        if len(available_engines) == 1:
            engine_info = f"使用引擎: {available_engines[0].name}"
        
        self.message_user(
            request, 
            f"已提交 {len(queryset)} 个镜像到后台队列拉取\n"
            f"稍后可刷新页面查看拉取结果",
            level='success'
        )
    
    pull_images.short_description = "拉取镜像"
    
    
    
    def changelist_view(self, request, extra_context=None):
        """
        重写列表视图，在页面加载时触发异步检查任务
        """
        from django.core.cache import cache
        from datetime import datetime, timedelta
        
        # 使用缓存锁避免频繁触发任务
        lock_key = 'docker_images_batch_check_lock'
        last_check_key = 'docker_images_last_batch_check'
        
        # 检查上次检查时间
        last_check_time = cache.get(last_check_key)
        current_time = datetime.now()
        
        # 如果超过3分钟没有检查，触发异步任务
        should_trigger = False
        if last_check_time is None:
            should_trigger = True
        else:
            time_diff = (current_time - last_check_time).total_seconds()
            if time_diff > 180:  # 3分钟
                should_trigger = True
        
        # 尝试获取锁并触发任务
        if should_trigger:
            # 使用 set nx (not exists) 来实现分布式锁
            lock_acquired = cache.add(lock_key, 'locked', timeout=60)
            
            if lock_acquired:
                try:
                    from easytask.tasks import batch_check_images_status
                    
                    # 异步触发任务
                    batch_check_images_status.delay()
                    
                    # 更新最后检查时间
                    cache.set(last_check_key, current_time, timeout=600)
                    
                    # 添加成功消息
                    
                    
                except Exception as e:
                    logger.error(f"触发批量检查任务失败，请检查Celery配置")
                finally:
                    # 释放锁
                    cache.delete(lock_key)
        
        return super().changelist_view(request, extra_context)
    
    class Media:
        """添加自定义 JavaScript"""
        js = ('admin/js/refresh_image_status.js',)





@admin.register(ContainerEngineConfig)
class ContainerEngineConfigAdmin(admin.ModelAdmin):
    """容器引擎配置管理 - 单例模式"""
    
    # 隐藏列表页，直接显示编辑页
    def has_add_permission(self, request):
        # 只允许一条记录，禁止添加
        return False
    
    def has_delete_permission(self, request, obj=None):
        # 不允许删除
        return False
    
    def changelist_view(self, request, extra_context=None):
        """列表视图重定向到编辑页"""
        obj, created = ContainerEngineConfig.objects.get_or_create(pk=1)
        from django.shortcuts import redirect
        from django.urls import reverse
        return redirect(reverse('admin:container_containerengineconfig_change', args=[obj.pk]))
    
    fieldsets = (
        ('🌐 容器生命周期配置', {
            'fields': ('container_expiry_hours',),
            'description': '<div style="background: #e7f3ff; padding: 10px; border-left: 4px solid #2196F3; margin-bottom: 10px; width: 50%;">容器创建后的有效期设置</div>'
        }),
        ('👥 容器并发限制', {
            'fields': (
                'max_containers_per_user',
                'max_containers_per_challenge',
                'max_containers_per_team',
                'max_concurrent_creates',
            ),
            'description': '<div style="background: #fff3cd; padding: 10px; border-left: 4px solid #ffc107; margin-bottom: 10px; width: 50%;">防止资源耗尽的并发限制</div>'
        }),
        ('🚦 令牌桶限流配置', {
            'fields': (
                'token_bucket_max',
                'token_bucket_refill_rate',
            ),
            'description': '<div style="background: #d4edda; padding: 10px; border-left: 4px solid #28a745; margin-bottom: 10px; width: 50%;">基于令牌桶的限流策略（第二层防护）</div>'
        }),
        ('🐋 Docker引擎配置', {
            'fields': (
                'docker_pool_min_size',
                'docker_pool_max_size',
                'docker_max_usage_threshold',
                'docker_image_pull_timeout',
            ),
            'description': '<div style="background: #d1ecf1; padding: 10px; border-left: 4px solid #17a2b8; margin-bottom: 10px; width: 50%;">Docker引擎连接池和资源阈值</div>'
        }),
        ('☸️ K8s节点资源阈值', {
            'fields': (
                'k8s_node_memory_threshold',
                'k8s_node_cpu_threshold',
            ),
            'description': '<div style="background: #f8d7da; padding: 10px; border-left: 4px solid #dc3545; margin-bottom: 10px; width: 50%;">单个节点的资源使用率限制</div>'
        }),
        ('☸️ K8s集群资源阈值', {
            'fields': (
                'k8s_cluster_memory_threshold',
                'k8s_cluster_cpu_threshold',
                'k8s_max_usage_threshold',
            ),
            'description': '<div style="background: #f8d7da; padding: 10px; border-left: 4px solid #dc3545; margin-bottom: 10px; width: 50%;">整个集群的资源使用率限制</div>'
        }),
        ('⚙️ K8s原子预占配置', {
            'fields': (
                'k8s_node_reservation_timeout',
                'k8s_node_cache_timeout',
            ),
            'description': '<div style="background: #e2e3e5; padding: 10px; border-left: 4px solid #6c757d; margin-bottom: 10px; width: 50%;">节点资源预占和缓存策略</div>'
        }),
        ('🎯 K8s资源策略', {
            'fields': (
                'k8s_requests_ratio',
                'k8s_use_max_node_capacity',
            ),
            'description': '<div style="background: #e2e3e5; padding: 10px; border-left: 4px solid #6c757d; margin-bottom: 10px; width: 50%;">K8s资源调度策略</div>'
        }),
        ('🔗 K8s API连接池配置', {
            'fields': (
                'k8s_connection_pool_maxsize',
                'k8s_connection_pool_block',
            ),
            'description': '<div style="background: #cfe2ff; padding: 10px; border-left: 4px solid #0d6efd; margin-bottom: 10px; width: 50%;">K8s API连接池配置，影响高并发性能</div>'
        }),
        ('📝 操作记录', {
            'fields': ('updated_at', 'updated_by'),
            'classes': ('collapse',),
        }),
    )
    
    readonly_fields = ['updated_at', 'updated_by']
    
    def save_model(self, request, obj, form, change):
        """保存时记录操作人"""
        obj.updated_by = request.user
        super().save_model(request, obj, form, change)
        
    


@admin.register(NetworkTopologyConfig)
class NetworkTopologyConfigAdmin(admin.ModelAdmin):
    list_display = ('name', 'nodes_count', 'author', 'created_at', 'is_active', 'editor_link')
    list_filter = ('is_active', 'created_at', 'author')
    search_fields = ('name', 'description')
    readonly_fields = ('created_at', 'updated_at', 'author')
    
    fieldsets = (
        (None, {
            'fields': ('name', 'description', 'is_active')
        }),
        ('创建信息', {
            'fields': ('author', 'created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    
    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if not request.user.is_superuser:
            qs = qs.filter(author=request.user)
        return qs
    
    def save_model(self, request, obj, form, change):
        if not obj.pk:
            obj.author = request.user
        super().save_model(request, obj, form, change)
    
    def nodes_count(self, obj):
        """获取拓扑节点数（使用模型方法以确保格式兼容）"""
        return obj.get_node_count()
    nodes_count.short_description = '节点数'
    
    def editor_link(self, obj):
        if obj.pk:
            from django.urls import reverse
            from django.utils.html import format_html
            url = reverse('container:topology_editor', args=[obj.pk])
            return format_html('<a href="{}" target="_blank">🎨 可视化编排</a>', url)
        return "-"
    editor_link.short_description = '操作'


