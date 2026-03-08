from django.contrib import admin
from django.urls import path, reverse
from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import HttpResponse
from practice.models import PC_Challenge, Tag, SolveRecord, SolvedFlag
from public.models import CTFUser
from container.models import DockerImage
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
import json

@admin.register(PC_Challenge)
class PC_ChallengeAdmin(admin.ModelAdmin):
    list_display = ('title', 'category', 'points', 'solves','flag_config_display', 'writeup_status_display' ,'is_active', 'author','updated_at')
    list_filter = ('category', 'difficulty', 'is_active')
    search_fields = ('title', 'description', 'author__username','tags__name')
    readonly_fields = ('solves', 'first_blood_time', 'allocated_coins', 'created_at', 'updated_at', 'purchased_count', 'flag_points_preview', 'flag_count_display')
    filter_horizontal = ('tags', 'writeup_purchased_by')
    # list_editable = ('docker_image', 'static_files', 'is_active')  # 已禁用：外键下拉框会导致严重的性能问题
    list_editable = ('is_active',)  # 只保留 is_active，外键字段请进入编辑页面修改
    
    # 自定义按钮
    actions = ['export_template_action', 'bulk_create_action', 'import_excel_action', 'auto_match_images_action']
    
    # 性能优化设置
    list_per_page = 50  # 每页显示50条
    list_max_show_all = 200  # 最多显示200条"显示全部"
    show_full_result_count = False  # 不显示精确的总数（大数据集时提升性能）
    
    def flag_config_display(self, obj):
        """显示flag配置状态"""
        if obj.flag_count > 1:
            points_list = obj.get_flag_points_list()
            points_str = ','.join(map(str, points_list))
            return f"🎯 {obj.flag_count}段 [{points_str}]"
        return f"🏁 单flag"
    flag_config_display.short_description = 'Flag配置'
    
    def writeup_status_display(self, obj):
        """显示题解状态"""
        if obj.writeup_is_public:
            return "✅ 公开"
        elif obj.hint:
            return f"🔒 {obj.writeup_cost}金币"
        return "无"
    writeup_status_display.short_description = '题解状态'
    
    def purchased_count(self, obj):
        """显示已购买用户数"""
        # 使用 annotate 预计算的值
        return getattr(obj, 'purchased_count_cache', obj.writeup_purchased_by.count())
    purchased_count.short_description = '已购买人数'
    
    def get_queryset(self, request):
        """优化查询性能，避免N+1问题"""
        from django.db.models import Count
        qs = super().get_queryset(request)
        # 预加载关联对象
        qs = qs.select_related('author', 'docker_image', 'static_files', 'first_blood_user')
        # 预加载多对多关系
        qs = qs.prefetch_related('tags', 'writeup_purchased_by')
        # 添加聚合字段，预计算已购买人数
        qs = qs.annotate(purchased_count_cache=Count('writeup_purchased_by'))
        return qs
    
    def get_form(self, request, obj=None, **kwargs):
        """设置表单初始值"""
        form = super().get_form(request, obj, **kwargs)
        # 创建新题目时，默认设置作者为当前用户
        if not obj and 'author' in form.base_fields:
            form.base_fields['author'].initial = request.user
        return form
    
    def save_model(self, request, obj, form, change):
        """保存模型时自动设置作者"""
        # 如果是新建题目且未设置作者，自动设置为当前用户
        if not change and not obj.author_id:
            obj.author = request.user
        super().save_model(request, obj, form, change)
    
    def flag_points_preview(self, obj):
        """预览flag分数分配"""
        points_list = obj.get_flag_points_list()
        total = sum(points_list)
        preview = " + ".join([f"Flag#{i+1}({p}分)" for i, p in enumerate(points_list)])
        
        if obj.flag_points and isinstance(obj.flag_points, list) and len(obj.flag_points) > 0:
            status = "✅ 自定义配置"
        else:
            status = "⚙️ 自动平均分配"
        
        return f"{status}\n{preview}\n总计: {total}分"
    flag_points_preview.short_description = 'Flag分数预览'
    
    def flag_count_display(self, obj):
        """静态flag时显示自动检测的数量"""
        if obj.flag_type == 'STATIC' and obj.flag_template:
            flags = [f.strip() for f in obj.flag_template.split(',') if f.strip()]
            return f"{len(flags)} 个 (自动检测)"
        return obj.flag_count
    flag_count_display.short_description = 'Flag数量'

    
    def get_fieldsets(self, request, obj=None):
        """根据flag_type动态调整显示的字段"""
        # 基本fieldsets结构
        fieldsets = [
            ('基本信息', {
                'fields': (
                    'title', 'description', 'category', 'difficulty', 
                    'points', 'coins', 'reward_coins', 'author'
                )
            }),
        ]
        
        # Flag配置区域 - 包含所有字段，由JavaScript控制显示/隐藏
        flag_fields = {
            'fields': ('flag_type', 'flag_template', 'flag_count', 'flag_points', 'flag_points_preview')
        }
        
        fieldsets.append(('Flag配置', flag_fields))
        
        # 其他fieldsets
        fieldsets.extend([
            ('部署资源', {
                'fields': ('static_files', 'static_file_url', 'docker_image', 'network_topology_config'),
                'classes': ('collapse',),
                'description': '根据部署类型选择相应的资源'
            }),
            ('题解配置', {
                'fields': ('hint', 'writeup_is_public', 'writeup_cost', 'writeup_purchased_by'),
                'description': '题目解析配置：勾选"题解是否公开"后所有用户可免费查看，否则需要消耗金币购买'
            }),
            ('其他信息', {
                'fields': ('tags', 'is_active', 'is_top','is_disable','is_member')
            }),
            ('统计信息', {
                'fields': ('solves', 'first_blood_user','first_blood_time', 'allocated_coins', 'purchased_count', 'created_at', 'updated_at'),
                'classes': ('collapse',)
            })
        ])
        
        return fieldsets
    
    # SimpleUI 自定义按钮 - 链接类型
    def export_template_action(self, request, queryset):
        """下载Excel模板按钮"""
        pass
    
    export_template_action.short_description = '下载Excel模板'
    export_template_action.icon = 'el-icon-download'
    export_template_action.type = 'success'
    export_template_action.action_type = 1  # 链接跳转
    export_template_action.action_url = '/adminx/practice/pc_challenge/export-template/'
    
    def bulk_create_action(self, request, queryset):
        """批量创建题目按钮"""
        pass
    
    bulk_create_action.short_description = '批量创建题目'
    bulk_create_action.icon = 'el-icon-plus'
    bulk_create_action.type = 'primary'
    bulk_create_action.action_type = 1  # 链接跳转
    bulk_create_action.action_url = '/adminx/practice/pc_challenge/bulk-create/'
    
    def import_excel_action(self, request, queryset):
        """批量导入题目按钮"""
        pass
    
    import_excel_action.short_description = '批量导入题目'
    import_excel_action.icon = 'el-icon-upload2'
    import_excel_action.type = 'warning'
    import_excel_action.action_type = 1  # 链接跳转
    import_excel_action.action_url = '/adminx/practice/pc_challenge/import-excel/'
    
    def auto_match_images_action(self, request, queryset):
        """智能匹配镜像按钮"""
        pass
    
    auto_match_images_action.short_description = '智能匹配镜像'
    auto_match_images_action.icon = 'el-icon-connection'
    auto_match_images_action.type = 'danger'
    auto_match_images_action.action_type = 1  # 链接跳转
    auto_match_images_action.action_url = '/adminx/practice/pc_challenge/auto-match-images/'
    
    def get_urls(self):
        """添加自定义URL"""
        urls = super().get_urls()
        custom_urls = [
            path('import-excel/', self.admin_site.admin_view(self.import_excel_view), name='practice_pc_challenge_import_excel'),
            path('export-template/', self.admin_site.admin_view(self.export_template_view), name='practice_pc_challenge_export_template'),
            path('bulk-create/', self.admin_site.admin_view(self.bulk_create_view), name='practice_pc_challenge_bulk_create'),
            path('auto-match-images/', self.admin_site.admin_view(self.auto_match_images_view), name='practice_pc_challenge_auto_match_images'),
        ]
        return custom_urls + urls
    
    def import_excel_view(self, request):
        """处理Excel导入"""
        # 权限检查
        if not self.has_add_permission(request):
            from django.core.exceptions import PermissionDenied
            raise PermissionDenied("您没有添加题目的权限")
        
        # GET请求显示上传页面
        if request.method != 'POST':
            context = dict(
                self.admin_site.each_context(request),
                title='批量导入题目',
            )
            return render(request, 'admin/practice/pc_challenge/import_page.html', context)
        
        # POST请求处理上传
        if not request.FILES.get('excel_file'):
            self.message_user(request, "请选择要导入的Excel文件", messages.ERROR)
            return redirect('.')
        
        excel_file = request.FILES['excel_file']
        
        try:
            # 读取Excel文件
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            
            success_count = 0
            error_messages = []
            
            # 从第2行开始读取数据（第1行是表头）
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # 跳过空行
                    if not any(row):
                        continue
                    
                    # 解析行数据
                    (title, description, category, difficulty, points, 
                     flag_type, flag_template, flag_count, flag_points_str,
                     coins, reward_coins, hint, writeup_is_public, writeup_cost,
                     is_active, is_top, is_disable, is_member, static_file_url) = row[:19]
                    
                    # 数据验证和转换
                    if not title:
                        error_messages.append(f"第{row_num}行: 标题不能为空")
                        continue
                    
                    # 处理flag_points
                    flag_points = []
                    if flag_points_str:
                        try:
                            flag_points = json.loads(flag_points_str)
                        except:
                            try:
                                # 尝试解析为逗号分隔的数字列表
                                flag_points = [int(x.strip()) for x in str(flag_points_str).split(',') if x.strip()]
                            except:
                                flag_points = []
                    
                    # 处理布尔值
                    writeup_is_public = str(writeup_is_public).lower() in ['true', '1', 'yes', '是', True, 1]
                    is_active = str(is_active).lower() in ['true', '1', 'yes', '是', True, 1]
                    is_top = str(is_top).lower() in ['true', '1', 'yes', '是', True, 1]
                    is_disable = str(is_disable).lower() in ['true', '1', 'yes', '是', True, 1]
                    is_member = str(is_member).lower() in ['true', '1', 'yes', '是', True, 1]
                    
                    # 创建题目
                    challenge = PC_Challenge.objects.create(
                        title=str(title).strip(),
                        description=str(description or '').strip(),
                        category=str(category or 'Web').strip(),
                        difficulty=str(difficulty or 'Medium').strip(),
                        points=int(points or 100),
                        flag_type=str(flag_type or 'DYNAMIC').strip(),
                        flag_template=str(flag_template or '').strip() if flag_template else None,
                        flag_count=int(flag_count or 1),
                        flag_points=flag_points if flag_points else [],
                        coins=int(coins or 4),
                        reward_coins=int(reward_coins or 0),
                        hint=str(hint or '').strip() if hint else None,
                        writeup_is_public=writeup_is_public,
                        writeup_cost=int(writeup_cost or 1),
                        is_active=is_active,
                        is_top=is_top,
                        is_disable=is_disable,
                        is_member=is_member,
                        static_file_url=str(static_file_url).strip() if static_file_url else None,
                        author=request.user
                    )
                    
                    success_count += 1
                    
                except Exception as e:
                    error_messages.append(f"第{row_num}行: {str(e)}")
                    continue
            
            # 显示结果消息
            if success_count > 0:
                self.message_user(request, f"成功导入 {success_count} 条题目数据！", messages.SUCCESS)
            else:
                self.message_user(request, "未导入任何数据", messages.WARNING)
            
            if error_messages:
                for msg in error_messages[:10]:  # 最多显示10条错误
                    self.message_user(request, msg, messages.WARNING)
                if len(error_messages) > 10:
                    self.message_user(request, f"...还有 {len(error_messages) - 10} 条错误未显示", messages.WARNING)
            
        except Exception as e:
            self.message_user(request, f"导入失败: {str(e)}", messages.ERROR)
        
        from django.urls import reverse
        return redirect(reverse('admin:practice_pc_challenge_changelist'))
    
    def export_template_view(self, request):
        """导出Excel模板"""
        # 权限检查 - 只有有查看权限的用户才能下载模板
        if not self.has_view_permission(request):
            from django.core.exceptions import PermissionDenied
            raise PermissionDenied("您没有查看权限")
        
        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '题目导入模板'
        
        # 定义表头
        headers = [
            '题目标题*', '题目描述*', '题目类型', '难度', '分数',
            'Flag类型', 'Flag值', 'Flag数量', 'Flag分数配置',
            '金币', '奖励金币', '题解内容', '题解公开', '题解金币',
            '是否激活', '是否置顶', '是否全局启用', '是否会员题目', '静态文件URL'
        ]
        
        # 设置表头样式
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 添加示例数据
        example_data = [
            'Web渗透入门', 
            '一道简单的Web题目，考察SQL注入基础知识', 
            'Web', 
            'Easy', 
            100,
            'STATIC', 
            'flag{sql_injection_easy},flag{union_select}', 
            2,
            '[60, 40]',
            2, 
            5, 
            '# 解题思路\n\n1. 发现SQL注入点\n2. 使用union注入', 
            'false', 
            1,
            'true', 
            'false', 
            'true', 
            'false',
            'https://example.com/files/challenge.zip'
        ]
        
        for col_num, value in enumerate(example_data, 1):
            cell = ws.cell(row=2, column=col_num, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # 设置列宽
        column_widths = [20, 40, 15, 10, 10, 12, 40, 12, 20, 10, 12, 40, 12, 12, 12, 12, 15, 15, 35]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width
        
        # 添加说明sheet
        ws_info = wb.create_sheet('填写说明')
        info_texts = [
            ['字段说明', ''],
            ['题目标题*', '必填，题目的标题'],
            ['题目描述*', '必填，题目的详细描述，支持Markdown格式'],
            ['题目类型', f"可选：{', '.join([c[0] for c in PC_Challenge.CATEGORY_CHOICES[:10]])}等，默认Web"],
            ['难度', 'Easy(简单)/Medium(中等)/Hard(困难)，默认Medium'],
            ['分数', '题目总分，默认100'],
            ['Flag类型', 'STATIC(静态Flag)/DYNAMIC(动态Flag)，默认DYNAMIC'],
            ['Flag值', '静态Flag时填写具体flag值，多个用逗号分隔；动态Flag时可不填'],
            ['Flag数量', '题目包含几个flag，1-10，默认1'],
            ['Flag分数配置', '每个flag的分数，格式: [30,30,40] 或 30,30,40，总和必须等于总分'],
            ['金币', '解题所需金币，默认4'],
            ['奖励金币', '解题奖励的金币，默认0'],
            ['题解内容', '题目的解题思路，支持Markdown格式'],
            ['题解公开', 'true/false，是否公开题解，默认false'],
            ['题解金币', '查看题解所需金币，默认1'],
            ['是否激活', 'true/false，是否激活该题目，默认true'],
            ['是否置顶', 'true/false，是否置顶显示，默认false'],
            ['是否全局启用', 'true/false，是否全局启用，默认true'],
            ['是否会员题目', 'true/false，是否仅会员可见，默认false'],
            ['静态文件URL', '静态文件的URL地址，如果填写了URL地址，则优先使用URL地址，否则使用静态文件'],
            ['', ''],
            ['注意事项', ''],
            ['1', '带*号的字段为必填项'],
            ['2', '导入时会自动设置作者为当前登录用户'],
            ['3', 'Flag分数配置的总和必须等于题目总分'],
            ['4', '静态Flag会自动根据逗号分隔检测Flag数量'],
            ['5', '布尔值字段可以使用: true/false, 1/0, yes/no, 是/否'],
        ]
        
        for row_num, (field, desc) in enumerate(info_texts, 1):
            ws_info.cell(row=row_num, column=1, value=field).font = Font(bold=True)
            ws_info.cell(row=row_num, column=2, value=desc)
        
        ws_info.column_dimensions['A'].width = 20
        ws_info.column_dimensions['B'].width = 60
        
        # 生成响应
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="challenge_import_template.xlsx"'
        
        return response
    
    def bulk_create_view(self, request):
        """批量创建题目"""
        # 权限检查
        if not self.has_add_permission(request):
            from django.core.exceptions import PermissionDenied
            raise PermissionDenied("您没有添加题目的权限")
        
        # GET请求显示创建表单
        if request.method != 'POST':
            from practice.models import Tag
            context = dict(
                self.admin_site.each_context(request),
                title='批量创建题目',
                categories=PC_Challenge.CATEGORY_CHOICES,
                tags=Tag.objects.all().order_by('name'),
            )
            return render(request, 'admin/practice/pc_challenge/bulk_create.html', context)
        
        # POST请求处理批量创建
        try:
            category = request.POST.get('category')
            count = int(request.POST.get('count', 10))
            name_prefix = request.POST.get('name_prefix', '').strip()
            difficulty = request.POST.get('difficulty', 'Medium')
            points = int(request.POST.get('points', 100))
            coins = int(request.POST.get('coins', 4))
            flag_type = request.POST.get('flag_type', 'DYNAMIC')
            description_template = request.POST.get('description_template', '').strip()
            is_active = request.POST.get('is_active') == 'on'
            selected_tag_ids = request.POST.getlist('tags')  # 获取选中的标签ID列表
            
            # 验证参数
            if not category:
                self.message_user(request, "请选择题目类型", messages.ERROR)
                return redirect('.')
            
            if count < 1 or count > 50:
                self.message_user(request, "创建数量必须在1-50之间", messages.ERROR)
                return redirect('.')
            
            # 确定名称前缀
            if not name_prefix:
                name_prefix = category
            
            # 查找已存在的题目，确定起始编号
            existing_titles = set(PC_Challenge.objects.filter(
                title__startswith=name_prefix
            ).values_list('title', flat=True))
            
            # 找到可用的起始编号
            start_num = 1
            while f"{name_prefix}{start_num}" in existing_titles:
                start_num += 1
            
            # 批量创建题目
            success_count = 0
            created_titles = []
            
            for i in range(count):
                num = start_num + i
                title = f"{name_prefix}{num}"
                
                # 生成描述
                if description_template:
                    description = description_template.replace('{{ category }}', category)
                    description = description.replace('{{ difficulty }}', difficulty)
                    description = description.replace('{{ number }}', str(num))
                else:
                    description = f"这是一道{category}类型的题目，难度为{difficulty}。"
                
                # 创建题目
                challenge = PC_Challenge.objects.create(
                    title=title,
                    description=description,
                    category=category,
                    difficulty=difficulty,
                    points=points,
                    coins=coins,
                    flag_type=flag_type,
                    flag_template=f"flag{{{title.lower()}}}" if flag_type == 'STATIC' else None,
                    flag_count=1,
                    is_active=is_active,
                    author=request.user
                )
                
                # 添加学习岛标签
                if selected_tag_ids:
                    challenge.tags.set(selected_tag_ids)
                
                success_count += 1
                created_titles.append(title)
            
            # 显示成功消息
            titles_preview = ', '.join(created_titles[:5])
            if len(created_titles) > 5:
                titles_preview += f' ... 等{len(created_titles)}道题目'
            
            self.message_user(
                request, 
                f"成功创建 {success_count} 道题目！({titles_preview})", 
                messages.SUCCESS
            )
            
        except ValueError as e:
            self.message_user(request, f"参数错误: {str(e)}", messages.ERROR)
            return redirect('.')
        except Exception as e:
            self.message_user(request, f"创建失败: {str(e)}", messages.ERROR)
            return redirect('.')
        
        return redirect(reverse('admin:practice_pc_challenge_changelist'))
    
    def auto_match_images_view(self, request):
        """智能匹配镜像并创建题目"""
        # 权限检查
        if not self.has_add_permission(request):
            from django.core.exceptions import PermissionDenied
            raise PermissionDenied("您没有添加题目的权限")
        
        # GET请求显示预览页面
        if request.method != 'POST':
            # 获取所有未被引用的镜像
            used_image_ids = PC_Challenge.objects.filter(
                docker_image__isnull=False
            ).values_list('docker_image_id', flat=True)
            
            unused_images = DockerImage.objects.exclude(
                id__in=used_image_ids
            ).order_by('category', 'name')
            
            # 按类型分组
            images_by_category = {}
            for image in unused_images:
                category = image.category
                if category not in images_by_category:
                    images_by_category[category] = []
                images_by_category[category].append(image)
            
            from practice.models import Tag
            context = dict(
                self.admin_site.each_context(request),
                title='智能匹配镜像创建题目',
                unused_images=unused_images,
                images_by_category=images_by_category,
                total_count=unused_images.count(),
                categories=PC_Challenge.CATEGORY_CHOICES,
                tags=Tag.objects.all().order_by('name'),
            )
            return render(request, 'admin/practice/pc_challenge/auto_match.html', context)
        
        # POST请求处理创建
        try:
            # 获取选中的镜像ID列表
            selected_image_ids = request.POST.getlist('selected_images')
            if not selected_image_ids:
                self.message_user(request, " 请至少选择一个镜像", messages.WARNING)
                return redirect('.')
            
            # 获取默认配置
            default_difficulty = request.POST.get('default_difficulty', 'Medium')
            default_points = int(request.POST.get('default_points', 100))
            default_coins = int(request.POST.get('default_coins', 10))
            default_flag_type = request.POST.get('default_flag_type', 'DYNAMIC')
            is_active = request.POST.get('is_active') == 'on'
            auto_generate_name = request.POST.get('auto_generate_name') == 'on'
            selected_tag_ids = request.POST.getlist('tags')  # 获取选中的标签ID列表
            
            # 批量创建题目
            success_count = 0
            created_titles = []
            errors = []
            
            for image_id in selected_image_ids:
                try:
                    image = DockerImage.objects.get(id=image_id)
                    
                    # 生成题目名称
                    if auto_generate_name:
                        # 使用完整的类型名称作为前缀（如"CVE复现1"、"Web1"等）
                        category_prefix = image.category
                        existing_nums = []
                        for challenge in PC_Challenge.objects.filter(category=image.category):
                            if challenge.title.startswith(category_prefix):
                                try:
                                    # 提取数字部分
                                    num_str = challenge.title.replace(category_prefix, '')
                                    if num_str.isdigit():
                                        existing_nums.append(int(num_str))
                                except ValueError:
                                    pass
                        
                        next_num = max(existing_nums) + 1 if existing_nums else 1
                        title = f"{category_prefix}{next_num}"
                    else:
                        # 使用镜像名称
                        title = f"{image.name.split('/')[-1]}-{image.tag}"
                    
                    # 检查重复
                    if PC_Challenge.objects.filter(title=title).exists():
                        title = f"{title}-{image.id}"
                    
                    # 生成描述
                    description = image.description or f"基于 {image.name}:{image.tag} 镜像的{image.category}题目"
                    
                    # 创建题目
                    challenge = PC_Challenge.objects.create(
                        title=title,
                        description=description,
                        category=image.category,
                        difficulty=default_difficulty,
                        points=default_points,
                        coins=default_coins,
                        flag_type=default_flag_type,
                        flag_count=1,
                        docker_image=image,
                        is_active=is_active,
                        author=request.user
                    )
                    
                    # 添加学习岛标签
                    if selected_tag_ids:
                        challenge.tags.set(selected_tag_ids)
                    
                    success_count += 1
                    created_titles.append(title)
                    
                except DockerImage.DoesNotExist:
                    errors.append(f"镜像ID {image_id} 不存在")
                except Exception as e:
                    errors.append(f"创建题目失败: {str(e)}")
            
            # 显示结果消息
            if success_count > 0:
                titles_preview = ', '.join(created_titles[:5])
                if len(created_titles) > 5:
                    titles_preview += f' ... 等{len(created_titles)}道题目'
                
                self.message_user(
                    request, 
                    f"成功创建 {success_count} 道题目！({titles_preview})", 
                    messages.SUCCESS
                )
            
            if errors:
                error_msg = '; '.join(errors[:3])
                if len(errors) > 3:
                    error_msg += f' ... 等{len(errors)}个错误'
                self.message_user(request, f"⚠️ {error_msg}", messages.WARNING)
            
        except Exception as e:
            self.message_user(request, f"❌ 操作失败: {str(e)}", messages.ERROR)
            return redirect('.')
        
        return redirect(reverse('admin:practice_pc_challenge_changelist'))
    
    def _get_category_abbr(self, category):
        """获取类型缩写"""
        abbr_map = {
            '签到': 'checkin',
            'Web': 'web',
            'Pwn': 'pwn',
            '逆向': 'reverse',
            '密码学': 'crypto',
            '杂项': 'misc',
            '数字取证': 'forensics',
            '内存取证': 'memory',
            '磁盘取证': 'disk',
            '流量分析': 'traffic',
            '日志分析': 'log',
            '移动安全': 'mobile',
            'Android': 'android',
            'iOS': 'ios',
            '物联网': 'iot',
            '区块链': 'blockchain',
            '智能合约': 'contract',
            '云安全': 'cloud',
            '容器安全': 'container',
            'AI安全': 'ai',
            '机器学习': 'ml',
            '开源情报': 'osint',
            '隐写术': 'stego',
            '编程': 'coding',
            '硬件安全': 'hardware',
            '无线电': 'radio',
            'CVE复现': 'cve',
            '渗透测试': 'pentest',
            '红队': 'redteam',
            '蓝队': 'blueteam',
            'AD域渗透': 'ad',
            '内网渗透': 'intranet',
            'Web3': 'web3',
            '元宇宙': 'metaverse',
            '游戏安全': 'game',
            '车联网': 'iov',
            '其他': 'other',
        }
        return abbr_map.get(category, 'challenge')

# CTFUser 和 CTFUserAdmin 已迁移到 public.models 和 public.admin
# 如需管理 CTFUser，请在 public.admin 中进行

@admin.register(Tag)
class TagAdmin(admin.ModelAdmin):
    list_display = ('name', 'id', 'slug',)
    search_fields = ('name',)

@admin.register(SolveRecord)
class SolveRecordAdmin(admin.ModelAdmin):
    list_display = ('user', 'challenge', 'solved_at')
    list_filter = ('challenge', 'solved_at')
    search_fields = ('user__user', 'challenge__title')
    date_hierarchy = 'solved_at'
    
    def get_queryset(self, request):
        return super().get_queryset(request).select_related('user', 'challenge')

@admin.register(SolvedFlag)
class SolvedFlagAdmin(admin.ModelAdmin):
    list_display = ('user_display', 'challenge', 'flag_display', 'points_earned', 'solved_at')
    list_filter = ('solved_at', 'flag_index', 'challenge__category', 'challenge__difficulty')
    search_fields = ('user__user__username', 'challenge__title')
    readonly_fields = ('user', 'challenge', 'flag_index', 'points_earned', 'solved_at', 'flag_hash', 'flag_progress_display')
    date_hierarchy = 'solved_at'
    
    def user_display(self, obj):
        """显示用户名"""
        return obj.user.user.username
    user_display.short_description = '用户'
    
    def flag_display(self, obj):
        """显示flag信息"""
        return f"🎯 Flag #{obj.flag_index + 1}/{obj.challenge.flag_count}"
    flag_display.short_description = 'Flag编号'
    
    def flag_progress_display(self, obj):
        """显示该用户在该题目的完整进度"""
        solved_flags = SolvedFlag.objects.filter(
            user=obj.user,
            challenge=obj.challenge
        ).order_by('flag_index')
        
        total_flags = obj.challenge.flag_count
        solved_count = solved_flags.count()
        total_points = sum(f.points_earned for f in solved_flags)
        
        progress = f"{solved_count}/{total_flags}"
        percentage = (solved_count / total_flags * 100) if total_flags > 0 else 0
        
        flags_status = []
        solved_indices = set(f.flag_index for f in solved_flags)
        for i in range(total_flags):
            if i in solved_indices:
                flags_status.append(f"✅ Flag #{i+1}")
            else:
                flags_status.append(f"⬜ Flag #{i+1}")
        
        return (
            f"进度: {progress} ({percentage:.1f}%)\n"
            f"已获得分数: {total_points}/{obj.challenge.points}\n"
            f"详细状态:\n" + "\n".join(flags_status)
        )
    flag_progress_display.short_description = '完整进度'
    
    def has_add_permission(self, request):
        """禁止手动添加（由系统自动创建）"""
        return False
    
    def has_change_permission(self, request, obj=None):
        """禁止修改（保持记录完整性）"""
        return False
    
    def get_queryset(self, request):
        return super().get_queryset(request).select_related('user__user', 'challenge')
    
    fieldsets = (
        ('基本信息', {
            'fields': ('user', 'challenge', 'flag_display')
        }),
        ('得分详情', {
            'fields': ('flag_index', 'points_earned', 'solved_at', 'flag_progress_display')
        }),
        ('技术信息', {
            'fields': ('flag_hash',),
            'classes': ('collapse',),
            'description': 'Flag的SHA256哈希值，用于验证唯一性'
        })
    )