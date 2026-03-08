"""
竞赛创建者视图
会员和管理员可以创建和管理自己的竞赛
"""
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.db import transaction, models
from django.db.models import Max
from django.core.paginator import Paginator
from django.utils import timezone
from django.core.cache import cache
from django.core.exceptions import ValidationError
from django.utils.html import escape
from quiz.models import Quiz, Question, Option, QuizQuestion, QuizRecord, QuizRegistration
from quiz.forms import QuizCreateForm, QuizEditForm, QuestionCreateForm, OptionFormSet
from comment.models import SystemNotification
import uuid
import logging
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

logger = logging.getLogger(__name__)


def check_creator_permission(user):
    """检查用户是否有创建竞赛的权限（会员或管理员）"""
    return user.is_staff or user.is_superuser or getattr(user, 'is_member', False)


def clear_quiz_cache(quiz_id):
    """清除竞赛相关的所有缓存"""
    cache_keys = [
        f'quiz_stats_{quiz_id}',
        f'quiz_registration_stats_{quiz_id}',
        f'quiz_leaderboard_{quiz_id}_10',
        f'quiz_leaderboard_{quiz_id}_50',
        f'quiz_leaderboard_{quiz_id}_100',
    ]
    cache.delete_many(cache_keys)
    # 清除列表缓存
    for status in ['', 'upcoming', 'ongoing', 'ended']:
        cache.delete(f'quiz_list_{status}_')


def _create_quiz_registration_notification(user, registration, quiz, status):
    """创建知识竞赛报名审核通知"""
    try:
        if status == 'approved':
            # 审核通过通知
            content = f'''
                <div class="notification-content">
                    
                    <p><strong>竞赛名称：</strong>{escape(quiz.title)}</p>
                    <p><strong>报名时间：</strong>{registration.created_at.strftime('%Y-%m-%d %H:%M')}</p>
                    <p><strong>审核时间：</strong>{timezone.now().strftime('%Y-%m-%d %H:%M')}</p>
                    <div class="mt-3 p-3 bg-light rounded">
                        <p class="mb-2"><i class="fa fa-info-circle text-primary mr-2"></i><strong>温馨提示：</strong></p>
                        <ul class="mb-0 pl-3 small">
                            <li>您现在可以参加该竞赛了</li>
                            <li>请在竞赛时间内完成答题</li>
                            <li>答题过程中请遵守诚信原则</li>
                        </ul>
                    </div>
                    <div class="mt-3">
                        <a href="/quiz/{quiz.slug}/" class="btn btn-primary btn-sm">
                            <i class="fa fa-arrow-right mr-1"></i>前往竞赛
                        </a>
                    </div>
                </div>
            '''
            title = '✅ 竞赛报名审核通过'
        elif status == 'rejected':
            # 审核拒绝通知
            content = f'''
                <div class="notification-content">
                    <p><strong>竞赛名称：</strong>{escape(quiz.title)}</p>
                    <p><strong>报名时间：</strong>{registration.created_at.strftime('%Y-%m-%d %H:%M')}</p>
                    <p><strong>审核时间：</strong>{timezone.now().strftime('%Y-%m-%d %H:%M')}</p>
                    <div class="mt-3 p-3 bg-light rounded">
                        <p class="mb-2"><i class="fa fa-info-circle text-warning mr-2"></i><strong>说明：</strong></p>
                        <ul class="mb-0 pl-3 small">
                            <li>您的报名申请未通过审核</li>
                            <li>如有疑问，请联系竞赛管理员</li>
                            <li>您可以查看其他可参加的竞赛</li>
                        </ul>
                    </div>
                    <div class="mt-3">
                        <a href="/quiz/" class="btn btn-secondary btn-sm">
                            <i class="fa fa-list mr-1"></i>查看其他竞赛
                        </a>
                    </div>
                </div>
            '''
            title = '❌ 竞赛报名审核未通过'
        else:
            return
        
        # 创建系统通知
        notification = SystemNotification.objects.create(
            title=title,
            content=content
        )
        notification.get_p.add(user)
        
        logger.info(f"Created registration notification for user {user.username}, quiz {quiz.slug}, status {status}")
        
    except Exception as e:
        logger.error(f"创建报名通知失败: {e}")
        # 通知发送失败不影响审核流程
        pass


@login_required
def my_quizzes(request):
    """我创建的竞赛列表"""
    if not check_creator_permission(request.user):
        messages.error(request, '您没有权限访问此页面，仅会员和管理员可用')
        return redirect('quiz:quiz_list')
    
    # 获取用户创建的竞赛
    quizzes = Quiz.objects.filter(creator=request.user).order_by('-created_at')
    
    # 为每个竞赛添加统计信息
    for quiz in quizzes:
        quiz.total_participants = QuizRecord.objects.filter(
            quiz=quiz,
            status__in=['completed', 'timeout']
        ).values('user').distinct().count()
        
        quiz.total_attempts = QuizRecord.objects.filter(
            quiz=quiz,
            status__in=['completed', 'timeout']
        ).count()
    
    # 分页
    paginator = Paginator(quizzes, 10)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'quizzes': page_obj,
        'page_obj': page_obj,
    }
    return render(request, 'quiz/creator/my_quizzes.html', context)


@login_required
@require_http_methods(["GET", "POST"])
def create_quiz(request):
    """创建新竞赛"""
    if not check_creator_permission(request.user):
        messages.error(request, '您没有权限创建竞赛，仅会员和管理员可用')
        return redirect('quiz:quiz_list')
    
    if request.method == 'POST':
        # 验证验证码
        captcha_key = request.POST.get('captcha_key')
        captcha_value = request.POST.get('captcha_value', '').strip().lower()
        
        if not captcha_key or not captcha_value:
            messages.error(request, '请输入验证码')
            form = QuizCreateForm()
            return render(request, 'quiz/creator/create_quiz.html', {'form': form})
        
        cached_captcha = cache.get(f'quiz_create_captcha_{captcha_key}')
        if not cached_captcha or cached_captcha.lower() != captcha_value:
            messages.error(request, '验证码错误或已过期')
            form = QuizCreateForm(request.POST, request.FILES)
            return render(request, 'quiz/creator/create_quiz.html', {'form': form})
        
        # 删除已使用的验证码
        cache.delete(f'quiz_create_captcha_{captcha_key}')
        
        # 使用 Form 处理数据
        form = QuizCreateForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                quiz = form.save(commit=False)
                quiz.creator = request.user
                quiz.is_active = False  # 新创建的竞赛默认不激活
                quiz.save()
                
                messages.success(request, f'竞赛"{quiz.title}"创建成功！请继续添加题目。')
                return redirect('quiz:creator_edit_quiz', quiz_slug=quiz.slug)
                
            except Exception as e:
                messages.error(request, f'创建失败：{str(e)}')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{form.fields[field].label}: {error}')
    else:
        # GET 请求，显示创建表单
        form = QuizCreateForm()
    
    return render(request, 'quiz/creator/create_quiz.html', {'form': form})


@login_required
def get_create_captcha(request):
    """获取创建竞赛的验证码"""
    if not check_creator_permission(request.user):
        return JsonResponse({'success': False, 'message': '无权限'})
    
    from public.utils import generate_captcha, generate_captcha_image
    
    captcha_text = generate_captcha()
    captcha_image = generate_captcha_image(captcha_text)  # 返回完整的 data URI
    captcha_key = str(uuid.uuid4())
    
    # 存储到 Redis，5分钟过期
    cache.set(f'quiz_create_captcha_{captcha_key}', captcha_text, 300)
    
    return JsonResponse({
        'success': True,
        'captcha_key': captcha_key,
        'captcha_image': captcha_image  # 返回完整的 data URI
    })


@login_required
def edit_quiz(request, quiz_slug):
    """编辑竞赛"""
    quiz = get_object_or_404(Quiz, slug=quiz_slug)
    
    # 检查权限
    if quiz.creator != request.user and not request.user.is_staff:
        messages.error(request, '您没有权限编辑此竞赛')
        return redirect('quiz:quiz_detail', quiz_slug=quiz_slug)
    
    if request.method == 'POST':
        form = QuizEditForm(request.POST, instance=quiz)
        if form.is_valid():
            form.save()
            # 清除相关缓存
            clear_quiz_cache(quiz.id)
            messages.success(request, '竞赛信息已更新')
            return redirect('quiz:creator_edit_quiz', quiz_slug=quiz.slug)
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{form.fields[field].label}: {error}')
    else:
        form = QuizEditForm(instance=quiz)
    
    # 获取竞赛的题目（分页）
    questions_list = quiz.quiz_questions.select_related('question').prefetch_related('question__options').order_by('order')
    
    # 分页处理
    paginator = Paginator(questions_list, 20)  # 每页20道题
    page_number = request.GET.get('page', 1)
    try:
        questions = paginator.get_page(page_number)
    except:
        questions = paginator.get_page(1)
    
    context = {
        'quiz': quiz,
        'form': form,
        'questions': questions,
        'total_questions': questions_list.count(),
    }
    return render(request, 'quiz/creator/edit_quiz.html', context)


@login_required
@require_http_methods(["POST"])
def delete_quiz(request, quiz_slug):
    """删除竞赛"""
    quiz = get_object_or_404(Quiz, slug=quiz_slug)
    
    # 检查权限
    if quiz.creator != request.user and not request.user.is_staff:
        return JsonResponse({'success': False, 'message': '您没有权限删除此竞赛'})
    
    # 检查是否有答题记录
    if quiz.records.exists():
        return JsonResponse({'success': False, 'message': '该竞赛已有用户答题记录，无法删除'})
    
    quiz_title = quiz.title
    quiz_id = quiz.id
    quiz.delete()
    
    # 清除相关缓存
    clear_quiz_cache(quiz_id)
    
    return JsonResponse({'success': True, 'message': f'竞赛"{quiz_title}"已删除'})


@login_required
def quiz_statistics(request, quiz_slug):
    """竞赛数据统计（优化：使用缓存）"""
    quiz = get_object_or_404(Quiz, slug=quiz_slug)
    
    # 检查权限
    if quiz.creator != request.user and not request.user.is_staff:
        messages.error(request, '您没有权限查看此竞赛的统计数据')
        return redirect('quiz:quiz_detail', quiz_slug=quiz_slug)
    
    # 尝试从缓存获取统计数据
    cache_key = f'quiz_stats_{quiz.id}'
    cached_stats = cache.get(cache_key)
    
    if cached_stats:
        stats = cached_stats['stats']
        distribution = cached_stats['distribution']
        violation_stats = cached_stats['violation_stats']
    else:
        # 获取统计数据
        stats = quiz.get_statistics()
        distribution = quiz.get_score_distribution()
        
        # 获取违规记录统计（使用聚合优化）
        from django.db.models import Count, Q
        violation_data = quiz.records.aggregate(
            total_violations=Count('id', filter=Q(violation_count__gt=0)),
            cheating_count=Count('id', filter=Q(status='cheating'))
        )
        violation_stats = {
            'total_violations': violation_data['total_violations'],
            'cheating_count': violation_data['cheating_count'],
        }
        
        # 缓存统计数据 3 分钟
        cache.set(cache_key, {
            'stats': stats,
            'distribution': distribution,
            'violation_stats': violation_stats
        }, 180)
    
    # 排行榜使用独立缓存（获取全部数据）
    all_leaderboard = quiz.get_leaderboard(limit=None)
    
    # 为排行榜添加排名
    for idx, item in enumerate(all_leaderboard, start=1):
        item['rank'] = idx
    
    # 排行榜分页
    leaderboard_paginator = Paginator(all_leaderboard, 20)  # 每页20条
    leaderboard_page = request.GET.get('leaderboard_page', 1)
    leaderboard = leaderboard_paginator.get_page(leaderboard_page)
    
    # 获取最近的答题记录（不缓存，保持实时性）
    all_recent_records = quiz.records.filter(
        status__in=['completed', 'timeout']
    ).select_related('user').order_by('-submit_time')
    
    # 最近答题记录分页
    recent_paginator = Paginator(all_recent_records, 20)  # 每页20条
    recent_page = request.GET.get('recent_page', 1)
    recent_records = recent_paginator.get_page(recent_page)
    
    context = {
        'quiz': quiz,
        'stats': stats,
        'distribution': distribution,
        'leaderboard': leaderboard,
        'recent_records': recent_records,
        'violation_stats': violation_stats,
    }
    return render(request, 'quiz/creator/quiz_statistics.html', context)


@login_required
def add_question_page(request, quiz_slug):
    """添加题目页面"""
    quiz = get_object_or_404(Quiz, slug=quiz_slug)
    
    # 检查权限
    if quiz.creator != request.user and not request.user.is_staff:
        messages.error(request, '您没有权限为此竞赛添加题目')
        return redirect('quiz:quiz_detail', quiz_slug=quiz_slug)
    
    # 获取可用的题目（未添加到当前竞赛的）
    existing_question_ids = quiz.quiz_questions.values_list('question_id', flat=True)
    available_questions = Question.objects.filter(is_active=True).exclude(
        id__in=existing_question_ids
    ).order_by('-created_at')
    
    # 分页
    paginator = Paginator(available_questions, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'quiz': quiz,
        'questions': page_obj,
        'page_obj': page_obj,
    }
    return render(request, 'quiz/creator/add_question.html', context)


@login_required
@require_http_methods(["POST"])
def add_question_to_quiz(request, quiz_slug):
    """将题目添加到竞赛"""
    quiz = get_object_or_404(Quiz, slug=quiz_slug)
    
    # 检查权限
    if quiz.creator != request.user and not request.user.is_staff:
        return JsonResponse({'success': False, 'message': '您没有权限添加题目'})
    
    question_id = request.POST.get('question_id')
    if not question_id:
        return JsonResponse({'success': False, 'message': '缺少题目ID'})
    
    try:
        question = get_object_or_404(Question, id=question_id, is_active=True)
        
        # 检查是否已添加
        if quiz.quiz_questions.filter(question=question).exists():
            return JsonResponse({'success': False, 'message': '该题目已添加过'})
        
        # 获取当前最大顺序号
        max_order = quiz.quiz_questions.aggregate(max_order=Max('order'))['max_order'] or 0
        
        # 添加题目
        QuizQuestion.objects.create(
            quiz=quiz,
            question=question,
            order=max_order + 1
        )
        
        # 重新计算总分
        quiz.calculate_total_score()
        
        return JsonResponse({
            'success': True,
            'message': f'题目"{question.content[:30]}..."已添加'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
@require_http_methods(["POST"])
def remove_question_from_quiz(request, quiz_slug):
    """从竞赛中移除题目"""
    import json
    
    quiz = get_object_or_404(Quiz, slug=quiz_slug)
    
    # 检查权限
    if quiz.creator != request.user and not request.user.is_staff:
        return JsonResponse({'success': False, 'message': '您没有权限移除题目'})
    
    try:
        # 解析JSON数据
        data = json.loads(request.body)
        question_id = data.get('question_id')
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'success': False, 'message': '无效的请求数据'})
    
    if not question_id:
        return JsonResponse({'success': False, 'message': '缺少题目ID'})
    
    try:
        qq = quiz.quiz_questions.filter(question_id=question_id).first()
        if not qq:
            return JsonResponse({'success': False, 'message': '题目不在此竞赛中'})
        
        qq.delete()
        
        # 重新计算总分
        quiz.calculate_total_score()
        
        return JsonResponse({'success': True, 'message': '题目已移除'})
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
@require_http_methods(["POST"])
def toggle_quiz_status(request, quiz_slug):
    """切换竞赛激活状态"""
    quiz = get_object_or_404(Quiz, slug=quiz_slug)
    
    # 检查权限
    if quiz.creator != request.user and not request.user.is_staff:
        return JsonResponse({'success': False, 'message': '您没有权限修改竞赛状态'})
    
    # 检查是否有题目
    if not quiz.quiz_questions.exists():
        return JsonResponse({'success': False, 'message': '请先添加题目再激活竞赛'})
    
    quiz.is_active = not quiz.is_active
    quiz.save(update_fields=['is_active'])
    
    # 清除相关缓存
    clear_quiz_cache(quiz.id)
    
    status_text = '已激活' if quiz.is_active else '已停用'
    return JsonResponse({'success': True, 'message': f'竞赛{status_text}', 'is_active': quiz.is_active})


@login_required
@require_http_methods(["GET", "POST"])
def create_question(request, quiz_slug):
    """创建新题目并添加到竞赛"""
    quiz = get_object_or_404(Quiz, slug=quiz_slug)
    
    # 检查权限
    if not check_creator_permission(request.user):
        messages.error(request, '您没有权限创建题目')
        return redirect('quiz:quiz_list')
    
    if quiz.creator != request.user and not request.user.is_staff:
        messages.error(request, '您没有权限为此竞赛添加题目')
        return redirect('quiz:creator_edit_quiz', quiz_slug=quiz_slug)
    
    if request.method == 'POST':
        form = QuestionCreateForm(request.POST)
        
        # 获取题目类型
        question_type = request.POST.get('question_type', '')
        
        # 填空题和简答题不需要验证选项
        if question_type in ['fill_blank', 'essay']:
            if form.is_valid():
                try:
                    with transaction.atomic():
                        # 保存题目
                        question = form.save(commit=False)
                        question.created_by = request.user
                        question.save()
                        
                        # 验证标准答案
                        if not question.standard_answer:
                            raise ValidationError('填空题和简答题必须填写标准答案')
                        
                        # 将题目添加到竞赛
                        max_order = quiz.quiz_questions.aggregate(max_order=models.Max('order'))['max_order'] or 0
                        QuizQuestion.objects.create(
                            quiz=quiz,
                            question=question,
                            order=max_order + 1
                        )
                        
                        # 重新计算竞赛总分
                        quiz.calculate_total_score()
                        
                        messages.success(request, f'题目创建成功！')
                        return redirect('quiz:creator_edit_quiz', quiz_slug=quiz.slug)
                        
                except ValidationError as e:
                    if hasattr(e, 'message'):
                        messages.error(request, f'创建失败：{e.message}')
                    elif hasattr(e, 'messages'):
                        for msg in e.messages:
                            messages.error(request, f'创建失败：{msg}')
                    else:
                        messages.error(request, f'创建失败：{str(e)}')
                except Exception as e:
                    messages.error(request, f'创建失败：{str(e)}')
            else:
                # 显示表单错误
                for field, errors in form.errors.items():
                    for error in errors:
                        messages.error(request, f'{form.fields[field].label}: {error}')
        else:
            # 选择题和判断题需要验证选项
            formset = OptionFormSet(
                request.POST,
                queryset=Option.objects.none()
            )
            
            if form.is_valid() and formset.is_valid():
                try:
                    with transaction.atomic():
                        # 保存题目
                        question = form.save(commit=False)
                        question.created_by = request.user
                        question.save()
                        
                        # 保存选项
                        options = formset.save(commit=False)
                        correct_count = 0
                        
                        # 选项标签映射：0->A, 1->B, 2->C...
                        option_labels = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
                        
                        for idx, option in enumerate(options):
                            option.question = question
                            option.order = option_labels[idx] if idx < len(option_labels) else str(idx + 1)
                            if option.is_correct:
                                correct_count += 1
                            option.save()
                        
                        # 验证选项数量
                        if question.question_type in ['single', 'multiple'] and len(options) != 4:
                            raise ValidationError('选择题必须有4个选项')
                        
                        if question.question_type == 'judge' and len(options) != 2:
                            raise ValidationError('判断题必须有2个选项')
                        
                        # 验证正确答案数量
                        if correct_count == 0:
                            raise ValidationError('至少需要设置一个正确答案')
                        
                        if question.question_type == 'single' and correct_count != 1:
                            raise ValidationError('单项选择题只能有1个正确答案')
                        
                        if question.question_type == 'judge' and correct_count != 1:
                            raise ValidationError('判断题只能有1个正确答案')
                        
                        if question.question_type == 'multiple' and correct_count < 2:
                            raise ValidationError('多项选择题至少需要2个正确答案')
                        
                        # 将题目添加到竞赛
                        max_order = quiz.quiz_questions.aggregate(max_order=models.Max('order'))['max_order'] or 0
                        QuizQuestion.objects.create(
                            quiz=quiz,
                            question=question,
                            order=max_order + 1
                        )
                        
                        # 重新计算竞赛总分
                        quiz.calculate_total_score()
                        
                        messages.success(request, f'题目创建成功！')
                        return redirect('quiz:creator_edit_quiz', quiz_slug=quiz.slug)
                        
                except ValidationError as e:
                    # 处理 ValidationError，提取错误消息
                    if hasattr(e, 'message'):
                        messages.error(request, f'创建失败：{e.message}')
                    elif hasattr(e, 'messages'):
                        for msg in e.messages:
                            messages.error(request, f'创建失败：{msg}')
                    else:
                        messages.error(request, f'创建失败：{str(e)}')
                except Exception as e:
                    messages.error(request, f'创建失败：{str(e)}')
            else:
                # 显示表单错误
                for field, errors in form.errors.items():
                    for error in errors:
                        messages.error(request, f'{form.fields[field].label}: {error}')
                
                for form_errors in formset.errors:
                    for field, errors in form_errors.items():
                        for error in errors:
                            messages.error(request, f'选项错误: {error}')
    else:
        form = QuestionCreateForm()
        # 创建formset，显示4个空表单
        # 使用 initial 来指定4个空选项
        formset = OptionFormSet(
            queryset=Option.objects.none(),
            initial=[{'content': '', 'is_correct': False} for _ in range(2)]
        )
    
    context = {
        'quiz': quiz,
        'form': form,
        'formset': formset,
    }
    return render(request, 'quiz/creator/create_question.html', context)


@login_required
def quiz_registrations(request, quiz_slug):
    """竞赛报名管理"""
    quiz = get_object_or_404(Quiz, slug=quiz_slug)
    
    # 检查权限
    if quiz.creator != request.user and not request.user.is_staff:
        messages.error(request, '您没有权限查看此竞赛的报名记录')
        return redirect('quiz:quiz_detail', quiz_slug=quiz_slug)
    
    # 检查竞赛是否启用报名功能
    if not quiz.require_registration:
        messages.info(request, '该竞赛未启用报名功能')
        return redirect('quiz:quiz_statistics', quiz_slug=quiz_slug)
    
    # 获取报名记录
    registrations = QuizRegistration.objects.filter(
        quiz=quiz
    ).select_related('user').order_by('-created_at')
    
    # 状态筛选
    status_filter = request.GET.get('status', '')
    if status_filter:
        registrations = registrations.filter(status=status_filter)
    
    # 分页
    paginator = Paginator(registrations, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # 统计数据（使用缓存和聚合优化，减少数据库查询）
    cache_key = f'quiz_registration_stats_{quiz.id}'
    stats = cache.get(cache_key)
    
    if not stats:
        from django.db.models import Count, Q
        stats = QuizRegistration.objects.filter(quiz=quiz).aggregate(
            total_count=Count('id'),
            approved_count=Count('id', filter=Q(status='approved')),
            pending_count=Count('id', filter=Q(status='pending')),
            rejected_count=Count('id', filter=Q(status='rejected'))
        )
        # 缓存 5 分钟
        cache.set(cache_key, stats, 300)
    
    total_count = stats['total_count']
    approved_count = stats['approved_count']
    pending_count = stats['pending_count']
    rejected_count = stats['rejected_count']
    
    context = {
        'quiz': quiz,
        'registrations': page_obj,
        'page_obj': page_obj,
        'total_count': total_count,
        'approved_count': approved_count,
        'pending_count': pending_count,
        'rejected_count': rejected_count,
        'status_filter': status_filter,
    }
    return render(request, 'quiz/creator/registrations.html', context)


@login_required
@require_http_methods(["POST"])
def update_registration_status(request, quiz_slug, registration_id):
    """更新报名状态（审核）"""
    quiz = get_object_or_404(Quiz, slug=quiz_slug)
    
    # 检查权限
    if quiz.creator != request.user and not request.user.is_staff:
        return JsonResponse({
            'success': False,
            'message': '您没有权限执行此操作'
        }, status=403)
    
    registration = get_object_or_404(QuizRegistration, id=registration_id, quiz=quiz)
    new_status = request.POST.get('status', '')
    
    if new_status not in ['approved', 'rejected', 'pending']:
        return JsonResponse({
            'success': False,
            'message': '无效的状态'
        }, status=400)
    
    try:
        old_status = registration.status
        registration.status = new_status
        registration.save(update_fields=['status', 'updated_at'])
        
        # 清除报名统计缓存
        cache.delete(f'quiz_registration_stats_{quiz.id}')
        cache.delete(f'user_registration_{quiz.id}_{registration.user.id}')
        
        # 发送系统通知
        if new_status in ['approved', 'rejected'] and old_status != new_status:
            _create_quiz_registration_notification(registration.user, registration, quiz, new_status)
        
        # 记录操作日志
        logger.info(
            f"Registration status updated: quiz={quiz.slug}, "
            f"user={registration.user.username}, "
            f"operator={request.user.username}, "
            f"old_status={old_status}, new_status={new_status}"
        )
        
        return JsonResponse({
            'success': True,
            'message': f'状态已更新为：{registration.get_status_display()}'
        })
    except Exception as e:
        logger.error(f"更新报名状态失败: {e}")
        return JsonResponse({
            'success': False,
            'message': f'更新失败：{str(e)}'
        }, status=500)


@login_required
def question_detail(request, question_id):
    """获取题目详情（需要权限验证）"""
    from quiz.templatetags.quiz_tags import markdown as render_markdown
    
    question = get_object_or_404(Question, id=question_id)
    
    # 权限检查：只有以下情况可以查看题目详情
    # 1. 管理员或超级用户
    # 2. 题目所属竞赛的创建者
    has_permission = False
    
    if request.user.is_staff or request.user.is_superuser:
        has_permission = True
    else:
        # 检查用户是否是任何包含此题目的竞赛的创建者
        user_quizzes = Quiz.objects.filter(
            creator=request.user,
            questions=question
        )
        if user_quizzes.exists():
            has_permission = True
    
    if not has_permission:
        return JsonResponse({
            'success': False,
            'message': '您没有权限查看此题目详情'
        }, status=403)
    
    # 类型颜色映射
    type_colors = {
        'single': 'primary',
        'multiple': 'info',
        'judge': 'warning',
        'fill_blank': 'success',
        'essay': 'danger'
    }
    
    # 渲染 Markdown
    content_html = render_markdown(question.content) if question.content else ''
    explanation_html = render_markdown(question.explanation) if question.explanation else ''
    standard_answer_html = render_markdown(question.standard_answer) if question.standard_answer else ''
    
    # 构建返回数据
    data = {
        'success': True,
        'question': {
            'id': question.id,
            'question_type': question.question_type,
            'type': question.question_type,
            'type_display': question.get_question_type_display(),
            'type_color': type_colors.get(question.question_type, 'secondary'),
            'content': content_html,
            'score': question.score,
            'difficulty': question.difficulty,
            'difficulty_display': question.get_difficulty_display(),
            'category': question.category,
            'explanation': explanation_html,
            'standard_answer': standard_answer_html,
            'options': [
                {
                    'order': opt.order,
                    'content': render_markdown(opt.content) if opt.content else opt.content,
                    'is_correct': opt.is_correct
                }
                for opt in question.options.all().order_by('order')
            ]
        }
    }
    
    return JsonResponse(data)


@login_required
def download_import_template(request):
    """下载题目导入模板"""
    wb = Workbook()
    ws = wb.active
    ws.title = "题目导入模板"
    
    # 设置标题行
    headers = ['题目类型*', '题目内容*', '选项A', '选项B', '选项C', '选项D', '正确答案/标准答案*', '分数', '难度', '分类', '答案解析']
    ws.append(headers)
    
    # 设置标题样式
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 添加示例数据
    examples = [
        ['单选', 'HTTP协议默认端口是多少？', '80', '443', '8080', '3000', 'A', '5', '简单', '网络基础', 'HTTP协议的默认端口是80'],
        ['多选', '以下哪些是Python的数据类型？', 'list', 'array', 'dict', 'tuple', 'ACD', '10', '中等', 'Python基础', 'Python内置的数据类型包括list、dict和tuple'],
        ['判断', 'SQL注入是一种常见的Web安全漏洞', '正确', '错误', '', '', 'A', '3', '简单', 'Web安全', 'SQL注入是最常见的Web漏洞之一'],
        ['填空', 'Python中用于定义函数的关键字是____', '', '', '', '', 'def', '5', '简单', 'Python基础', '使用def关键字定义函数'],
        ['简答', '请简述SQL注入的原理和防御方法', '', '', '', '', 'SQL注入是通过在输入中插入恶意SQL代码来操纵数据库。防御方法包括：使用参数化查询、输入验证、最小权限原则等。', '15', '困难', 'Web安全', '主要考察对SQL注入攻击的理解和防御措施'],
    ]
    for example in examples:
        ws.append(example)
    
    # 设置列宽
    column_widths = [12, 50, 20, 20, 20, 20, 40, 8, 10, 15, 50]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    # 添加说明sheet
    ws_info = wb.create_sheet("导入说明")
    instructions = [
        ['字段说明', ''],
        ['题目类型*', '必填。可填写：单选、多选、判断、填空、简答'],
        ['题目内容*', '必填。题目的问题描述'],
        ['选项A', '选择题和判断题必填，填空题和简答题不需要'],
        ['选项B', '选择题和判断题必填，填空题和简答题不需要'],
        ['选项C', '选择题选填，判断题、填空题、简答题不需要'],
        ['选项D', '选择题选填，判断题、填空题、简答题不需要'],
        ['正确答案/标准答案*', '必填。客观题填A/B/C/D等；主观题填写标准答案（用于批改参考）'],
        ['分数', '选填。默认1分'],
        ['难度', '选填。可填：简单、中等、困难。默认中等'],
        ['分类', '选填。题目分类标签'],
        ['答案解析', '选填。答案解析说明'],
        ['', ''],
        ['题型说明', ''],
        ['客观题（自动判分）', '单选题、多选题、判断题'],
        ['主观题（人工批改）', '填空题、简答题'],
        ['', ''],
        ['注意事项', ''],
        ['1. 带*号的字段为必填项', ''],
        ['2. 题目类型必须是：单选、多选、判断、填空、简答 之一', ''],
        ['3. 单选和多选题必须有4个选项（A、B、C、D）', ''],
        ['4. 判断题只需要2个选项（A、B），通常填"正确"和"错误"', ''],
        ['5. 填空题和简答题不需要选项，直接填写标准答案', ''],
        ['6. 正确答案/标准答案格式：', ''],
        ['   - 单选题：A 或 B 或 C 或 D', ''],
        ['   - 多选题：AB 或 AC 或 BCD 等（多个答案连写）', ''],
        ['   - 判断题：A 或 B', ''],
        ['   - 填空题：标准答案文本（如：def）', ''],
        ['   - 简答题：参考答案文本（用于批改参考）', ''],
        ['7. 主观题需要人工批改，系统会自动将其分配给阅卷人', ''],
        ['8. 导入前请删除示例数据', ''],
    ]
    for row in instructions:
        ws_info.append(row)
    
    ws_info.column_dimensions['A'].width = 25
    ws_info.column_dimensions['B'].width = 60
    
    # 设置说明标题样式
    for cell in ws_info['A']:
        if cell.value in ['字段说明', '题型说明', '注意事项']:
            cell.font = Font(bold=True, size=14)
            cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="quiz_questions_template.xlsx"'
    wb.save(response)
    return response


@login_required
@require_http_methods(["POST"])
def import_questions(request, slug):
    """批量导入题目"""
    quiz = get_object_or_404(Quiz, slug=slug)
    
    # 权限检查
    if quiz.creator != request.user and not request.user.is_staff:
        return JsonResponse({
            'success': False,
            'message': '您没有权限导入题目到此竞赛'
        }, status=403)
    
    excel_file = request.FILES.get('excel_file')
    if not excel_file:
        return JsonResponse({
            'success': False,
            'message': '请选择要上传的Excel文件',
            'errors': []
        })
    
    if not excel_file.name.endswith(('.xlsx', '.xls')):
        return JsonResponse({
            'success': False,
            'message': '只支持Excel格式文件（.xlsx 或 .xls）',
            'errors': []
        })
    
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        success_count = 0
        error_list = []
        
        with transaction.atomic():
            # 获取当前最大顺序号
            max_order = QuizQuestion.objects.filter(quiz=quiz).aggregate(
                max_order=Max('order')
            )['max_order'] or 0
            
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # 解析行数据
                    question_type_str = str(row[0]).strip() if row[0] else ''
                    content = str(row[1]).strip() if row[1] else ''
                    option_a = str(row[2]).strip() if row[2] else ''
                    option_b = str(row[3]).strip() if row[3] else ''
                    option_c = str(row[4]).strip() if row[4] else ''
                    option_d = str(row[5]).strip() if row[5] else ''
                    correct_answer = str(row[6]).strip().upper() if row[6] else ''
                    score = float(row[7]) if row[7] else 1.0
                    difficulty_str = str(row[8]).strip() if row[8] else '中等'
                    category = str(row[9]).strip() if row[9] else ''
                    explanation = str(row[10]).strip() if row[10] else ''
                    
                    # 跳过空行
                    if not content:
                        continue
                    
                    # 映射题目类型
                    type_mapping = {
                        '单选': 'single',
                        '多选': 'multiple',
                        '判断': 'judge',
                        '填空': 'fill_blank',
                        '简答': 'essay'
                    }
                    question_type = type_mapping.get(question_type_str)
                    if not question_type:
                        error_list.append(f'第{row_num}行：题目类型无效（{question_type_str}）')
                        continue
                    
                    # 映射难度
                    difficulty_mapping = {'简单': 'easy', '中等': 'medium', '困难': 'hard'}
                    difficulty = difficulty_mapping.get(difficulty_str, 'medium')
                    
                    # 创建题目
                    question = Question.objects.create(
                        question_type=question_type,
                        content=content,
                        explanation=explanation,
                        score=score,
                        difficulty=difficulty,
                        category=category if category else None,
                        is_active=True
                    )
                    
                    # 根据题型创建选项或标准答案
                    if question_type in ['fill_blank', 'essay']:
                        # 填空题和简答题：设置标准答案
                        if not correct_answer:
                            error_list.append(f'第{row_num}行：填空题和简答题必须填写标准答案')
                            question.delete()
                            continue
                        question.standard_answer = correct_answer
                        question.save(update_fields=['standard_answer'])
                    else:
                        # 客观题：创建选项
                        options_data = []
                        if question_type in ['single', 'multiple']:
                            if not all([option_a, option_b, option_c, option_d]):
                                error_list.append(f'第{row_num}行：选择题必须有4个选项')
                                question.delete()
                                continue
                            options_data = [
                                ('A', option_a),
                                ('B', option_b),
                                ('C', option_c),
                                ('D', option_d)
                            ]
                        elif question_type == 'judge':
                            if not all([option_a, option_b]):
                                error_list.append(f'第{row_num}行：判断题必须有2个选项')
                                question.delete()
                                continue
                            options_data = [
                                ('A', option_a),
                                ('B', option_b)
                            ]
                        
                        # 创建选项
                        for order, opt_content in options_data:
                            is_correct = order in correct_answer
                            Option.objects.create(
                                question=question,
                                order=order,
                                content=opt_content,
                                is_correct=is_correct
                            )
                        
                        # 验证客观题选项
                        errors = question.validate_options()
                        if errors:
                            error_list.append(f'第{row_num}行：{"; ".join(errors)}')
                            question.delete()
                            continue
                    
                    # 添加到竞赛
                    max_order += 1
                    QuizQuestion.objects.create(
                        quiz=quiz,
                        question=question,
                        order=max_order
                    )
                    
                    success_count += 1
                    
                except Exception as e:
                    error_list.append(f'第{row_num}行：导入失败 - {str(e)}')
                    continue
        
        # 重新计算总分
        if success_count > 0:
            quiz.calculate_total_score()
        
        # 清除缓存
        clear_quiz_cache(quiz.id)
        
        # 返回结果
        if success_count > 0:
            return JsonResponse({
                'success': True,
                'message': f'成功导入 {success_count} 道题目' + (f'，{len(error_list)} 个错误' if error_list else ''),
                'count': success_count,
                'errors': error_list
            })
        else:
            return JsonResponse({
                'success': False,
                'message': '导入失败，没有成功导入任何题目',
                'errors': error_list
            })
            
    except Exception as e:
        logger.error(f"导入题目失败: {e}")
        return JsonResponse({
            'success': False,
            'message': f'文件解析失败：{str(e)}',
            'errors': []
        })
